import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# CoA segment 2 = Location/Department (matches the existing CoA Segments sample)
columns = [
    ("Kitchen",    2, "Back-of-house kitchen production department"),
    ("Front POS",  2, "Front-of-house point-of-sale / dining room department"),
    ("Bar",        2, "Bar and lounge department"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EOD Columns"

HDR_BG = "2B6B35"; HDR_FG = "FFFFFF"; ALT = "F0F7F0"
header_fill  = PatternFill("solid", fgColor=HDR_BG)
header_font  = Font(name="Arial", bold=True, color=HDR_FG, size=11)
header_align = Alignment(horizontal="center", vertical="center")
alt_fill     = PatternFill("solid", fgColor=ALT)
data_font    = Font(name="Arial", size=10)
thin         = Border(
    left=Side(style="thin", color="C8DFC8"), right=Side(style="thin", color="C8DFC8"),
    top=Side(style="thin", color="C8DFC8"),  bottom=Side(style="thin", color="C8DFC8"),
)

headers    = ["Name", "CoA Segment", "Description"]
col_widths = [18,     12,            50]

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = header_align; cell.border = thin
    ws.column_dimensions[cell.column_letter].width = w
ws.row_dimensions[1].height = 20

for row_idx, (name, seg, desc) in enumerate(columns, start=2):
    fill = alt_fill if row_idx % 2 == 0 else PatternFill()
    for col, val in enumerate([name, seg, desc], start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = data_font; cell.fill = fill; cell.border = thin
        cell.alignment = Alignment(vertical="center",
                                   horizontal="center" if col == 2 else "left")

ws.freeze_panes = "A2"
out = r"C:\ClaudeOutput\Culinaire\SampleData\EOD Columns.xlsx"
wb.save(out)
print(f"Saved: {out}  ({len(columns)} columns)")
