from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill("solid", fgColor="2B6B35")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="F0F7F0")
NORMAL_FONT = Font(name="Arial", size=10)
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(bottom=THIN)

methods = [
    ("FedEx Ground",       "FedEx ground delivery — typically 1-5 business days within contiguous US"),
    ("FedEx Express",      "FedEx overnight or 2-day air express service"),
    ("UPS Ground",         "UPS ground delivery — typically 1-5 business days, cost-effective for heavy shipments"),
    ("UPS Next Day Air",   "UPS guaranteed next-business-day delivery by end of day"),
    ("USPS Priority Mail", "USPS Priority Mail — 1-3 business days with free tracking and up to $100 insurance"),
]

wb = Workbook()
ws = wb.active
ws.title = "ShippingMethods"

headers = ["Name", "Description"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")

for r, (name, desc) in enumerate(methods, 2):
    fill = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate([name, desc], 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = NORMAL_FONT
        if fill:
            cell.fill = fill
        cell.border = BORDER

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 18
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 68

wb.save(r"C:\ClaudeOutput\Culinaire\SampleData\ShippingMethods.xlsx")
print("ShippingMethods.xlsx written.")
