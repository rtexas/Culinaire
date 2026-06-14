from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, name="Arial", size=10, color="FFFFFF")
DATA_FONT   = Font(name="Arial", size=10)

def style_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = DATA_FONT
    for i, h in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        max_len = max(len(str(h)), *(len(str(r[i-1])) for r in rows if i-1 < len(r)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

DIR = r"C:\ClaudeOutput\Culinaire\SampleData"

# EodSections
wb = Workbook(); ws = wb.active; ws.title = "Sections"
style_sheet(ws,
    ["Name", "Description", "Multiplier"],
    [
        ["Sales",        "Gross revenue items - added to column totals",                          1],
        ["Deductions",   "Comps / voids / discounts - subtracted from column totals",            -1],
        ["Liabilities",  "Sales tax and other pass-through amounts - excluded from net total",    0],
        ["Dispositions", "Payment method collections - subtracted from totals to reconcile",     -1],
    ])
wb.save(os.path.join(DIR, "EodSections.xlsx"))
print("EodSections.xlsx saved")

# EodColumns
wb = Workbook(); ws = wb.active; ws.title = "Columns"
style_sheet(ws,
    ["Name", "Description", "CoA Segment", "Segment Value"],
    [
        ["Bar",       "Bar / Beverage revenue center",               1, "BAR"],
        ["Kitchen",   "Kitchen / Food production revenue center",    1, "KITCH"],
        ["Front POS", "Front-of-house / Dining room revenue center", 1, "FPOS"],
        ["Catering",  "Catering and banquet revenue center",         1, "CATR"],
        ["Delivery",  "Delivery and takeout revenue center",         1, "DLVR"],
    ])
wb.save(os.path.join(DIR, "EodColumns.xlsx"))
print("EodColumns.xlsx saved")

# EodRows
wb = Workbook(); ws = wb.active; ws.title = "Rows"
style_sheet(ws,
    ["Name", "Description", "Section"],
    [
        ["Food Sales",              "Total food revenue",                              "Sales"],
        ["Liquor Sales",            "Distilled spirits revenue",                       "Sales"],
        ["Beer Sales",              "Draught and bottled beer revenue",                "Sales"],
        ["Wine Sales",              "Wine by the glass and bottle revenue",            "Sales"],
        ["Non-Alcoholic Beverages", "Soda / juice / coffee / tea revenue",             "Sales"],
        ["Catering Sales",          "Catering and banquet event revenue",              "Sales"],
        ["Delivery Sales",          "Third-party and in-house delivery revenue",       "Sales"],
        ["Takeout Sales",           "To-go / takeout order revenue",                   "Sales"],
        ["Comps",                   "Complimentary items issued to guests",            "Deductions"],
        ["Employee Meals",          "Discounted or free meals provided to staff",      "Deductions"],
        ["Voids",                   "Voided items removed from checks",                "Deductions"],
        ["Discounts",               "Promotional and loyalty discounts applied",       "Deductions"],
        ["Sales Tax",               "State and local sales tax collected",             "Liabilities"],
        ["Cash",                    "Cash payments collected",                         "Dispositions"],
        ["Checks",                  "Check payments collected",                        "Dispositions"],
        ["Visa",                    "Visa credit and debit card payments",             "Dispositions"],
        ["Amex",                    "American Express card payments",                  "Dispositions"],
        ["Discover",                "Discover card payments",                          "Dispositions"],
    ])
wb.save(os.path.join(DIR, "EodRows.xlsx"))
print("EodRows.xlsx saved")

# ChartOfAccounts
wb = Workbook(); ws = wb.active; ws.title = "Chart Of Accounts"
accounts = [
    ["Bar - Food Sales",                   "BAR-001-4000",  "Food items sold through the bar",                    "Revenue",   "Sales",      True],
    ["Bar - Liquor Sales",                 "BAR-001-4100",  "Distilled spirits sold through the bar",             "Revenue",   "Sales",      True],
    ["Bar - Beer Sales",                   "BAR-001-4200",  "Draught and bottled beer sold through the bar",      "Revenue",   "Sales",      True],
    ["Bar - Wine Sales",                   "BAR-001-4300",  "Wine sold through the bar",                          "Revenue",   "Sales",      True],
    ["Bar - Non-Alcoholic Beverages",      "BAR-001-4400",  "Soda / juice / coffee sold through the bar",         "Revenue",   "Sales",      True],
    ["Bar - Comps",                        "BAR-001-4800",  "Complimentary items issued at the bar",              "Revenue",   "Deductions", True],
    ["Bar - Employee Meals",               "BAR-001-4850",  "Staff meals charged through the bar",                "Revenue",   "Deductions", True],
    ["Bar - Voids",                        "BAR-001-4900",  "Voided transactions at the bar",                     "Revenue",   "Deductions", True],
    ["Bar - Discounts",                    "BAR-001-4950",  "Promotional discounts applied at the bar",           "Revenue",   "Deductions", True],
    ["Bar - Sales Tax",                    "BAR-001-2100",  "Sales tax collected at the bar",                     "Liability", "Tax",        True],
    ["Kitchen - Food Sales",               "KITCH-001-4000","Food items produced by the kitchen",                 "Revenue",   "Sales",      True],
    ["Kitchen - Catering Sales",           "KITCH-001-4500","Catering and event food produced by kitchen",        "Revenue",   "Sales",      True],
    ["Kitchen - Comps",                    "KITCH-001-4800","Complimentary food items from the kitchen",          "Revenue",   "Deductions", True],
    ["Kitchen - Employee Meals",           "KITCH-001-4850","Staff meals produced by the kitchen",               "Revenue",   "Deductions", True],
    ["Kitchen - Voids",                    "KITCH-001-4900","Voided kitchen items",                               "Revenue",   "Deductions", True],
    ["Kitchen - Discounts",                "KITCH-001-4950","Discounts applied to kitchen items",                 "Revenue",   "Deductions", True],
    ["Kitchen - Sales Tax",                "KITCH-001-2100","Sales tax on kitchen food items",                    "Liability", "Tax",        True],
    ["Front POS - Food Sales",             "FPOS-001-4000", "Food items sold through the dining room POS",        "Revenue",   "Sales",      True],
    ["Front POS - Liquor Sales",           "FPOS-001-4100", "Spirits sold through the dining room POS",           "Revenue",   "Sales",      True],
    ["Front POS - Beer Sales",             "FPOS-001-4200", "Beer sold through the dining room POS",              "Revenue",   "Sales",      True],
    ["Front POS - Wine Sales",             "FPOS-001-4300", "Wine sold through the dining room POS",              "Revenue",   "Sales",      True],
    ["Front POS - Non-Alcoholic Beverages","FPOS-001-4400", "Non-alcoholic drinks through dining room POS",       "Revenue",   "Sales",      True],
    ["Front POS - Comps",                  "FPOS-001-4800", "Complimentary items in the dining room",             "Revenue",   "Deductions", True],
    ["Front POS - Employee Meals",         "FPOS-001-4850", "Staff meals through the dining room POS",            "Revenue",   "Deductions", True],
    ["Front POS - Voids",                  "FPOS-001-4900", "Voided dining room transactions",                    "Revenue",   "Deductions", True],
    ["Front POS - Discounts",              "FPOS-001-4950", "Discounts applied in the dining room",               "Revenue",   "Deductions", True],
    ["Front POS - Sales Tax",              "FPOS-001-2100", "Sales tax collected through dining room POS",        "Liability", "Tax",        True],
    ["Catering - Catering Sales",          "CATR-001-4500", "Catering and banquet event revenue",                 "Revenue",   "Sales",      True],
    ["Catering - Delivery Sales",          "CATR-001-4600", "Off-site catering delivery revenue",                 "Revenue",   "Sales",      True],
    ["Catering - Comps",                   "CATR-001-4800", "Complimentary items for catering events",            "Revenue",   "Deductions", True],
    ["Catering - Voids",                   "CATR-001-4900", "Voided catering transactions",                       "Revenue",   "Deductions", True],
    ["Catering - Discounts",               "CATR-001-4950", "Discounts applied to catering orders",               "Revenue",   "Deductions", True],
    ["Catering - Sales Tax",               "CATR-001-2100", "Sales tax on catering revenue",                      "Liability", "Tax",        True],
    ["Delivery - Delivery Sales",          "DLVR-001-4600", "Third-party and in-house delivery revenue",          "Revenue",   "Sales",      True],
    ["Delivery - Takeout Sales",           "DLVR-001-4700", "To-go / takeout order revenue",                      "Revenue",   "Sales",      True],
    ["Delivery - Comps",                   "DLVR-001-4800", "Complimentary items on delivery orders",             "Revenue",   "Deductions", True],
    ["Delivery - Voids",                   "DLVR-001-4900", "Voided delivery transactions",                       "Revenue",   "Deductions", True],
    ["Delivery - Discounts",               "DLVR-001-4950", "Discounts applied to delivery orders",               "Revenue",   "Deductions", True],
    ["Delivery - Sales Tax",               "DLVR-001-2100", "Sales tax collected on delivery orders",             "Liability", "Tax",        True],
    # ── Cash & Credit Card accounts by Dept-Location ──────────────────────
    # Dept-Location combos: POS→001,002 | BAR→001,002 | KITCH→001,002,003
    #   BAKERY→001 | CATR→003 | DELIV→004 | EVENT→001,003 | PATIO→001
    #   PRVT→001   | CORP→004
    ["Front POS - Main Restaurant - Cash",            "FPOS-001-1100",   "Cash and checks collected at Front POS",        "Asset", "Cash",        True],
    ["Front POS - Main Restaurant - Credit Card",     "FPOS-001-1300",   "Credit card receipts at Front POS",             "Asset", "Credit Card", True],
    ["Front POS - Bar & Lounge - Cash",               "FPOS-002-1100",   "Cash and checks collected at Front POS",        "Asset", "Cash",        True],
    ["Front POS - Bar & Lounge - Credit Card",        "FPOS-002-1300",   "Credit card receipts at Front POS",             "Asset", "Credit Card", True],
    ["Bar - Main Restaurant - Cash",                  "BAR-001-1100",   "Cash and checks collected at Bar",              "Asset", "Cash",        True],
    ["Bar - Main Restaurant - Credit Card",           "BAR-001-1300",   "Credit card receipts at Bar",                   "Asset", "Credit Card", True],
    ["Bar - Bar & Lounge - Cash",                     "BAR-002-1100",   "Cash and checks collected at Bar",              "Asset", "Cash",        True],
    ["Bar - Bar & Lounge - Credit Card",              "BAR-002-1300",   "Credit card receipts at Bar",                   "Asset", "Credit Card", True],
    ["Kitchen - Main Restaurant - Cash",              "KITCH-001-1100", "Cash and checks collected at Kitchen",          "Asset", "Cash",        True],
    ["Kitchen - Main Restaurant - Credit Card",       "KITCH-001-1300", "Credit card receipts at Kitchen",               "Asset", "Credit Card", True],
    ["Kitchen - Bar & Lounge - Cash",                 "KITCH-002-1100", "Cash and checks collected at Kitchen",          "Asset", "Cash",        True],
    ["Kitchen - Bar & Lounge - Credit Card",          "KITCH-002-1300", "Credit card receipts at Kitchen",               "Asset", "Credit Card", True],
    ["Kitchen - Catering - Cash",                     "KITCH-003-1100", "Cash and checks collected at Kitchen",          "Asset", "Cash",        True],
    ["Kitchen - Catering - Credit Card",              "KITCH-003-1300", "Credit card receipts at Kitchen",               "Asset", "Credit Card", True],
    ["Bakery - Main Restaurant - Cash",               "BAKERY-001-1100","Cash and checks collected at Bakery",           "Asset", "Cash",        True],
    ["Bakery - Main Restaurant - Credit Card",        "BAKERY-001-1300","Credit card receipts at Bakery",                "Asset", "Credit Card", True],
    ["Catering - Catering - Cash",                    "CATR-003-1100",  "Cash and checks collected at Catering",         "Asset", "Cash",        True],
    ["Catering - Catering - Credit Card",             "CATR-003-1300",  "Credit card receipts at Catering",              "Asset", "Credit Card", True],
    ["Delivery - Delivery - Cash",                    "DELIV-004-1100", "Cash and checks collected at Delivery",         "Asset", "Cash",        True],
    ["Delivery - Delivery - Credit Card",             "DELIV-004-1300", "Credit card receipts at Delivery",              "Asset", "Credit Card", True],
    ["Event Space - Main Restaurant - Cash",          "EVENT-001-1100", "Cash and checks collected at Event Space",      "Asset", "Cash",        True],
    ["Event Space - Main Restaurant - Credit Card",   "EVENT-001-1300", "Credit card receipts at Event Space",           "Asset", "Credit Card", True],
    ["Event Space - Catering - Cash",                 "EVENT-003-1100", "Cash and checks collected at Event Space",      "Asset", "Cash",        True],
    ["Event Space - Catering - Credit Card",          "EVENT-003-1300", "Credit card receipts at Event Space",           "Asset", "Credit Card", True],
    ["Patio - Main Restaurant - Cash",                "PATIO-001-1100", "Cash and checks collected at Patio",            "Asset", "Cash",        True],
    ["Patio - Main Restaurant - Credit Card",         "PATIO-001-1300", "Credit card receipts at Patio",                 "Asset", "Credit Card", True],
    ["Private Dining - Main Restaurant - Cash",       "PRVT-001-1100",  "Cash and checks collected at Private Dining",   "Asset", "Cash",        True],
    ["Private Dining - Main Restaurant - Credit Card","PRVT-001-1300",  "Credit card receipts at Private Dining",        "Asset", "Credit Card", True],
    ["Corporate - Delivery - Cash",                   "CORP-004-1100",  "Cash and checks collected at Corporate",        "Asset", "Cash",        True],
    ["Corporate - Delivery - Credit Card",            "CORP-004-1300",  "Credit card receipts at Corporate",             "Asset", "Credit Card", True],
    # ── Accounts Payable — one per location ───────────────────────────────
    ["Main Restaurant - Accounts Payable",            "001-2100",       "Accounts payable for Main Restaurant",          "Liability", "Accounts Payable", True],
    ["Bar & Lounge - Accounts Payable",               "002-2100",       "Accounts payable for Bar & Lounge",             "Liability", "Accounts Payable", True],
    ["Catering - Accounts Payable",                   "003-2100",       "Accounts payable for Catering location",        "Liability", "Accounts Payable", True],
    ["Delivery - Accounts Payable",                   "004-2100",       "Accounts payable for Delivery location",        "Liability", "Accounts Payable", True],
    # ── Department Expense Accounts (6000-6999) ───────────────────────────
    # Front POS — locs 001 (Main Restaurant), 002 (Bar & Lounge)
    ["Front POS - Main Restaurant - Dining Supplies", "FPOS-001-6200",  "Paper goods, napkins, and dining room supplies at Main Restaurant", "Expense", "Operating Expenses", True],
    ["Front POS - Bar & Lounge - Dining Supplies",    "FPOS-002-6200",  "Paper goods, napkins, and dining room supplies at Bar & Lounge",    "Expense", "Operating Expenses", True],
    # Bar — locs 001 (Main Restaurant), 002 (Bar & Lounge)
    ["Bar - Main Restaurant - Liquor Cost",           "BAR-001-6010",   "Cost of distilled spirits sold at Bar / Main Restaurant",           "Expense", "Cost of Sales",      True],
    ["Bar - Main Restaurant - Beer Cost",             "BAR-001-6020",   "Cost of beer sold at Bar / Main Restaurant",                        "Expense", "Cost of Sales",      True],
    ["Bar - Main Restaurant - Wine Cost",             "BAR-001-6030",   "Cost of wine sold at Bar / Main Restaurant",                        "Expense", "Cost of Sales",      True],
    ["Bar - Main Restaurant - Bar Supplies",          "BAR-001-6210",   "Bar supplies, garnishes, and smallwares at Main Restaurant",        "Expense", "Operating Expenses", True],
    ["Bar - Bar & Lounge - Liquor Cost",              "BAR-002-6010",   "Cost of distilled spirits sold at Bar & Lounge",                    "Expense", "Cost of Sales",      True],
    ["Bar - Bar & Lounge - Beer Cost",                "BAR-002-6020",   "Cost of beer sold at Bar & Lounge",                                 "Expense", "Cost of Sales",      True],
    ["Bar - Bar & Lounge - Wine Cost",                "BAR-002-6030",   "Cost of wine sold at Bar & Lounge",                                 "Expense", "Cost of Sales",      True],
    ["Bar - Bar & Lounge - Bar Supplies",             "BAR-002-6210",   "Bar supplies, garnishes, and smallwares at Bar & Lounge",           "Expense", "Operating Expenses", True],
    # Kitchen — locs 001, 002, 003
    ["Kitchen - Main Restaurant - Food Cost",         "KITCH-001-6000", "Cost of food produced by kitchen at Main Restaurant",               "Expense", "Cost of Sales",      True],
    ["Kitchen - Main Restaurant - Kitchen Supplies",  "KITCH-001-6200", "Kitchen supplies, chemicals, and smallwares at Main Restaurant",    "Expense", "Operating Expenses", True],
    ["Kitchen - Main Restaurant - Equipment Repair",  "KITCH-001-6400", "Kitchen equipment repair and maintenance at Main Restaurant",       "Expense", "Operating Expenses", True],
    ["Kitchen - Bar & Lounge - Food Cost",            "KITCH-002-6000", "Cost of food produced by kitchen at Bar & Lounge",                  "Expense", "Cost of Sales",      True],
    ["Kitchen - Bar & Lounge - Kitchen Supplies",     "KITCH-002-6200", "Kitchen supplies, chemicals, and smallwares at Bar & Lounge",       "Expense", "Operating Expenses", True],
    ["Kitchen - Bar & Lounge - Equipment Repair",     "KITCH-002-6400", "Kitchen equipment repair and maintenance at Bar & Lounge",          "Expense", "Operating Expenses", True],
    ["Kitchen - Catering - Food Cost",                "KITCH-003-6000", "Cost of food produced by kitchen at Catering",                      "Expense", "Cost of Sales",      True],
    ["Kitchen - Catering - Kitchen Supplies",         "KITCH-003-6200", "Kitchen supplies, chemicals, and smallwares at Catering",           "Expense", "Operating Expenses", True],
    ["Kitchen - Catering - Equipment Repair",         "KITCH-003-6400", "Kitchen equipment repair and maintenance at Catering",              "Expense", "Operating Expenses", True],
    # Bakery — loc 001
    ["Bakery - Main Restaurant - Food Cost",          "BAKERY-001-6000","Cost of bakery ingredients and pastry supplies",                    "Expense", "Cost of Sales",      True],
    ["Bakery - Main Restaurant - Bakery Supplies",    "BAKERY-001-6200","Bakery packaging, boxes, and production supplies",                  "Expense", "Operating Expenses", True],
    # Catering — loc 003
    ["Catering - Catering - Food Cost",               "CATR-003-6000",  "Cost of food for catering events",                                  "Expense", "Cost of Sales",      True],
    ["Catering - Catering - Supplies",                "CATR-003-6200",  "Catering supplies, serving equipment, and disposables",             "Expense", "Operating Expenses", True],
    ["Catering - Catering - Marketing",               "CATR-003-6500",  "Catering marketing, event promotion, and advertising",              "Expense", "Operating Expenses", True],
    # Delivery — loc 004
    ["Delivery - Delivery - Packaging",               "DELIV-004-6230", "Delivery packaging materials and containers",                       "Expense", "Operating Expenses", True],
    ["Delivery - Delivery - Supplies",                "DELIV-004-6200", "Delivery bags, insulated carriers, and supplies",                   "Expense", "Operating Expenses", True],
    # Event Space — locs 001, 003
    ["Event Space - Main Restaurant - Supplies",      "EVENT-001-6200", "Event supplies, linen, and setup materials at Main Restaurant",     "Expense", "Operating Expenses", True],
    ["Event Space - Main Restaurant - Marketing",     "EVENT-001-6500", "Event marketing and promotional expenses at Main Restaurant",       "Expense", "Operating Expenses", True],
    ["Event Space - Catering - Supplies",             "EVENT-003-6200", "Event supplies, linen, and setup materials at Catering",            "Expense", "Operating Expenses", True],
    ["Event Space - Catering - Marketing",            "EVENT-003-6500", "Event marketing and promotional expenses at Catering",              "Expense", "Operating Expenses", True],
    # Patio — loc 001
    ["Patio - Main Restaurant - Supplies",            "PATIO-001-6200", "Patio supplies, outdoor tableware, and umbrellas",                  "Expense", "Operating Expenses", True],
    # Private Dining — loc 001
    ["Private Dining - Main Restaurant - Supplies",   "PRVT-001-6200",  "Private dining supplies, special linen, and decor",                "Expense", "Operating Expenses", True],
    # Corporate — loc 004
    ["Corporate - Delivery - Software & Technology",  "CORP-004-6610",  "Software subscriptions, POS licenses, and technology costs",       "Expense", "Operating Expenses", True],
    ["Corporate - Delivery - Miscellaneous",          "CORP-004-6700",  "Miscellaneous corporate operating expenses",                        "Expense", "Operating Expenses", True],
    # ── Location Expense Accounts (utilities, occupancy, etc.) ────────────
    # Main Restaurant (loc 001)
    ["Main Restaurant - Food Expense",                "001-6050",       "Food purchases for Main Restaurant",                                "Expense", "Cost of Sales",      True],
    ["Main Restaurant - Credit Card Expense",         "001-6510",       "Credit card merchant fees and charges for Main Restaurant",        "Expense", "General & Admin",    True],
    ["Main Restaurant - Interest Expense",            "001-6520",       "Interest charges on credit cards and financing for Main Restaurant","Expense", "General & Admin",    True],
    ["Main Restaurant - Electric Utilities",          "001-6300",       "Electric utility expense for Main Restaurant",                      "Expense", "Occupancy",          True],
    ["Main Restaurant - Natural Gas",                 "001-6310",       "Natural gas utility expense for Main Restaurant",                   "Expense", "Occupancy",          True],
    ["Main Restaurant - Water & Sewer",               "001-6320",       "Water and sewer utility expense for Main Restaurant",               "Expense", "Occupancy",          True],
    ["Main Restaurant - Internet & Telephone",        "001-6330",       "Internet and telephone service for Main Restaurant",                "Expense", "Occupancy",          True],
    ["Main Restaurant - Building Maintenance",        "001-6340",       "Building maintenance and repairs for Main Restaurant",              "Expense", "Occupancy",          True],
    ["Main Restaurant - Pest Control",                "001-6420",       "Pest control service for Main Restaurant",                          "Expense", "Occupancy",          True],
    ["Main Restaurant - Landscaping",                 "001-6430",       "Landscaping and grounds maintenance for Main Restaurant",           "Expense", "Occupancy",          True],
    ["Main Restaurant - Insurance",                   "001-6620",       "Property and liability insurance for Main Restaurant",              "Expense", "General & Admin",    True],
    ["Main Restaurant - Professional Services",       "001-6630",       "Accounting, legal, and consulting fees for Main Restaurant",       "Expense", "General & Admin",    True],
    ["Main Restaurant - Rent & Occupancy",            "001-6810",       "Rent and base occupancy costs for Main Restaurant",                 "Expense", "Occupancy",          True],
    # Bar & Lounge (loc 002)
    ["Bar & Lounge - Food Expense",                   "002-6050",       "Food purchases for Bar & Lounge",                                   "Expense", "Cost of Sales",      True],
    ["Bar & Lounge - Credit Card Expense",            "002-6510",       "Credit card merchant fees and charges for Bar & Lounge",           "Expense", "General & Admin",    True],
    ["Bar & Lounge - Interest Expense",               "002-6520",       "Interest charges on credit cards and financing for Bar & Lounge",  "Expense", "General & Admin",    True],
    ["Bar & Lounge - Electric Utilities",             "002-6300",       "Electric utility expense for Bar & Lounge",                         "Expense", "Occupancy",          True],
    ["Bar & Lounge - Natural Gas",                    "002-6310",       "Natural gas utility expense for Bar & Lounge",                      "Expense", "Occupancy",          True],
    ["Bar & Lounge - Water & Sewer",                  "002-6320",       "Water and sewer utility expense for Bar & Lounge",                  "Expense", "Occupancy",          True],
    ["Bar & Lounge - Internet & Telephone",           "002-6330",       "Internet and telephone service for Bar & Lounge",                   "Expense", "Occupancy",          True],
    ["Bar & Lounge - Building Maintenance",           "002-6340",       "Building maintenance and repairs for Bar & Lounge",                 "Expense", "Occupancy",          True],
    ["Bar & Lounge - Pest Control",                   "002-6420",       "Pest control service for Bar & Lounge",                             "Expense", "Occupancy",          True],
    ["Bar & Lounge - Landscaping",                    "002-6430",       "Landscaping and grounds maintenance for Bar & Lounge",              "Expense", "Occupancy",          True],
    ["Bar & Lounge - Insurance",                      "002-6620",       "Property and liability insurance for Bar & Lounge",                 "Expense", "General & Admin",    True],
    ["Bar & Lounge - Professional Services",          "002-6630",       "Accounting, legal, and consulting fees for Bar & Lounge",          "Expense", "General & Admin",    True],
    ["Bar & Lounge - Rent & Occupancy",               "002-6810",       "Rent and base occupancy costs for Bar & Lounge",                   "Expense", "Occupancy",          True],
    # Catering (loc 003)
    ["Catering - Food Expense",                       "003-6050",       "Food purchases for Catering location",                              "Expense", "Cost of Sales",      True],
    ["Catering - Credit Card Expense",                "003-6510",       "Credit card merchant fees and charges for Catering location",      "Expense", "General & Admin",    True],
    ["Catering - Interest Expense",                   "003-6520",       "Interest charges on credit cards and financing for Catering",      "Expense", "General & Admin",    True],
    ["Catering - Electric Utilities",                 "003-6300",       "Electric utility expense for Catering location",                    "Expense", "Occupancy",          True],
    ["Catering - Natural Gas",                        "003-6310",       "Natural gas utility expense for Catering location",                 "Expense", "Occupancy",          True],
    ["Catering - Water & Sewer",                      "003-6320",       "Water and sewer utility expense for Catering location",             "Expense", "Occupancy",          True],
    ["Catering - Internet & Telephone",               "003-6330",       "Internet and telephone service for Catering location",              "Expense", "Occupancy",          True],
    ["Catering - Building Maintenance",               "003-6340",       "Building maintenance and repairs for Catering location",            "Expense", "Occupancy",          True],
    ["Catering - Pest Control",                       "003-6420",       "Pest control service for Catering location",                        "Expense", "Occupancy",          True],
    ["Catering - Landscaping",                        "003-6430",       "Landscaping and grounds maintenance for Catering location",         "Expense", "Occupancy",          True],
    ["Catering - Insurance",                          "003-6620",       "Property and liability insurance for Catering location",            "Expense", "General & Admin",    True],
    ["Catering - Professional Services",              "003-6630",       "Accounting, legal, and consulting fees for Catering location",     "Expense", "General & Admin",    True],
    ["Catering - Rent & Occupancy",                   "003-6810",       "Rent and base occupancy costs for Catering location",              "Expense", "Occupancy",          True],
    # Delivery (loc 004)
    ["Delivery - Food Expense",                       "004-6050",       "Food purchases for Delivery location",                              "Expense", "Cost of Sales",      True],
    ["Delivery - Credit Card Expense",                "004-6510",       "Credit card merchant fees and charges for Delivery location",      "Expense", "General & Admin",    True],
    ["Delivery - Interest Expense",                   "004-6520",       "Interest charges on credit cards and financing for Delivery",      "Expense", "General & Admin",    True],
    ["Delivery - Electric Utilities",                 "004-6300",       "Electric utility expense for Delivery location",                    "Expense", "Occupancy",          True],
    ["Delivery - Natural Gas",                        "004-6310",       "Natural gas utility expense for Delivery location",                 "Expense", "Occupancy",          True],
    ["Delivery - Water & Sewer",                      "004-6320",       "Water and sewer utility expense for Delivery location",             "Expense", "Occupancy",          True],
    ["Delivery - Internet & Telephone",               "004-6330",       "Internet and telephone service for Delivery location",              "Expense", "Occupancy",          True],
    ["Delivery - Building Maintenance",               "004-6340",       "Building maintenance and repairs for Delivery location",            "Expense", "Occupancy",          True],
    ["Delivery - Pest Control",                       "004-6420",       "Pest control service for Delivery location",                        "Expense", "Occupancy",          True],
    ["Delivery - Landscaping",                        "004-6430",       "Landscaping and grounds maintenance for Delivery location",         "Expense", "Occupancy",          True],
    ["Delivery - Insurance",                          "004-6620",       "Property and liability insurance for Delivery location",            "Expense", "General & Admin",    True],
    ["Delivery - Professional Services",              "004-6630",       "Accounting, legal, and consulting fees for Delivery location",     "Expense", "General & Admin",    True],
    ["Delivery - Rent & Occupancy",                   "004-6810",       "Rent and base occupancy costs for Delivery location",              "Expense", "Occupancy",          True],
]
style_sheet(ws,
    ["Account Name", "Account", "Account Description", "Account Category", "Account Type", "Active"],
    accounts)
wb.save(os.path.join(DIR, "ChartOfAccounts.xlsx"))
print("ChartOfAccounts.xlsx saved")

# Locations
wb = Workbook(); ws = wb.active; ws.title = "Locations"
style_sheet(ws,
    ["Code", "Name", "Description", "Segment Number", "Address1", "Address2", "City", "State", "Zip"],
    [
        ["001", "Main Restaurant",  "Primary dining location",   2, "1234 Westheimer Rd",   "",           "Houston", "TX", "77006"],
        ["002", "Bar & Lounge",     "Bar and lounge area",       2, "5678 Montrose Blvd",   "",           "Houston", "TX", "77006"],
        ["003", "Catering",         "Catering and events",       2, "910 Post Oak Blvd",    "Suite 100",  "Houston", "TX", "77056"],
        ["004", "Delivery",         "Delivery and takeout",      2, "2345 Richmond Ave",    "",           "Houston", "TX", "77098"],
    ])
wb.save(os.path.join(DIR, "Locations.xlsx"))
print("Locations.xlsx saved")

# Theme
def rgb(hex_str):
    h = hex_str.lstrip('#')
    return f"{int(h[0:2],16)},{int(h[2:4],16)},{int(h[4:6],16)}"

wb = Workbook(); ws = wb.active; ws.title = "Theme"
theme_defaults = [
    # Property,              Value,                                             Hex_Color,  RGB_Color
    ["Portal Name",          "Culinaire",                                       "",         ""],
    ["Tagline",              "Distinctive Dining & Hospitality Management",     "",         ""],
    ["Background Color",     "",  "#FFFFFF",  rgb("#FFFFFF")],
    ["Text Color",           "",  "#000000",  rgb("#000000")],
    ["Primary Color",        "",  "#2B6B35",  rgb("#2B6B35")],
    ["Accent Color",         "",  "#1A4A22",  rgb("#1A4A22")],
    ["Sidebar BG",           "",  "#1A4A22",  rgb("#1A4A22")],
    ["Sidebar Text",         "",  "#FFFFFF",  rgb("#FFFFFF")],
    ["Header BG",            "",  "#2B6B35",  rgb("#2B6B35")],
    ["Header Text",          "",  "#FFFFFF",  rgb("#FFFFFF")],
    ["Footer BG",            "",  "#1A4A22",  rgb("#1A4A22")],
    ["Footer Text",          "",  "#FFFFFF",  rgb("#FFFFFF")],
]
style_sheet(ws, ["Property", "Value", "Hex_Color", "RGB_Color"], theme_defaults)
wb.save(os.path.join(DIR, "Theme.xlsx"))
print("Theme.xlsx saved")

# Departments
wb = Workbook(); ws = wb.active; ws.title = "Departments"
departments = [
    ("POS",    "Front POS",      "Point of sale — front of house dining",         "1", "001"),
    ("POS",    "Front POS",      "Point of sale — front of house dining",         "1", "002"),
    ("BAR",    "Bar",            "Bar and beverage service area",                  "1", "001"),
    ("BAR",    "Bar",            "Bar and beverage service area",                  "1", "002"),
    ("KITCH",  "Kitchen",        "Back of house food preparation",                 "1", "001"),
    ("KITCH",  "Kitchen",        "Back of house food preparation",                 "1", "002"),
    ("KITCH",  "Kitchen",        "Back of house food preparation",                 "1", "003"),
    ("BAKERY", "Bakery",         "In-house bakery and pastry production",          "1", "001"),
    ("CATR",   "Catering",       "Off-site and event catering operations",         "1", "003"),
    ("DELIV",  "Delivery",       "Delivery and takeout order management",          "1", "004"),
    ("EVENT",  "Event Space",    "Private events and banquet room",                "1", "001"),
    ("EVENT",  "Event Space",    "Private events and banquet room",                "1", "003"),
    ("PATIO",  "Patio",          "Outdoor patio seating area",                     "1", "001"),
    ("PRVT",   "Private Dining", "Private dining room for groups",                 "1", "001"),
    ("CORP",   "Corporate",      "Corporate office and administrative functions",  "1", "004"),
]
headers_dept = ["Code", "Name", "Description", "IsActive", "LocationCode"]
ws.append(headers_dept)
for cell in ws[1]:
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
for row in departments:
    ws.append(list(row))
    for cell in ws[ws.max_row]:
        cell.font = DATA_FONT
col_widths = {"Code": 10, "Name": 18, "Description": 48, "IsActive": 10, "LocationCode": 14}
from openpyxl.utils import get_column_letter
for i, h in enumerate(headers_dept, 1):
    ws.column_dimensions[get_column_letter(i)].width = col_widths.get(h, 16)
wb.save(os.path.join(DIR, "Departments.xlsx"))
print("Departments.xlsx saved")

# Vendors
# DefaultPayablesAccount matches FullAccountString in ChartOfAccounts (e.g. "001-2100")
# Each vendor is assigned to the Main Restaurant AP account by default; change per location as needed.
wb = Workbook(); ws = wb.active; ws.title = "Vendors"
vendors = [
    ["SYSCO01",  "Sysco Corporation",                    "Broadline food distributor",                  "1390 Enclave Pkwy",     None,        None, "Houston",          "TX", "77077", "US", "001-2100"],
    ["USFDS01",  "US Foods",                             "National broadline distributor",               "9399 W Higgins Rd",     None,        None, "Rosemont",         "IL", "60018", "US", "001-2100"],
    ["PFGFT01",  "Performance Food Group",               "Broadline food and beverage distributor",      "12500 W Creek Pkwy",    None,        None, "Richmond",         "VA", "23238", "US", "001-2100"],
    ["GPFSC01",  "Gordon Food Service",                  "Broadline food distributor",                   "1300 Gezon Pkwy SW",    None,        None, "Wyoming",          "MI", "49509", "US", "001-2100"],
    ["REJFS01",  "Reinhart Foodservice",                 "Regional broadline distributor",               "6250 N Wilson Dr",      None,        None, "Milwaukee",        "WI", "53223", "US", "001-2100"],
    ["BENCO01",  "Ben E Keith Foods",                    "Southwest food distributor",                   "601 E 7th St",          None,        None, "Fort Worth",       "TX", "76102", "US", "001-2100"],
    ["DOLE01",   "Dole Food Company",                    "Fresh produce supplier",                       "1000 Corporate Ctr Dr", None,        None, "Monterey Park",    "CA", "91754", "US", "001-2100"],
    ["TAYFM01",  "Taylor Farms",                         "Fresh-cut produce and salads",                 "10 Becerra Rd",         None,        None, "Salinas",          "CA", "93905", "US", "001-2100"],
    ["CARGM01",  "Cargill Meat Solutions",               "Beef and protein supplier",                    "151 N Main St",         None,        None, "Wichita",          "KS", "67202", "US", "001-2100"],
    ["TYSON01",  "Tyson Foods",                          "Poultry and protein distributor",              "2200 Don Tyson Pkwy",   None,        None, "Springdale",       "AR", "72762", "US", "001-2100"],
    ["NBRKE01",  "Nebraska Beef",                        "Premium beef supplier",                        "605 S 3rd St",          None,        None, "Omaha",            "NE", "68102", "US", "001-2100"],
    ["LANDB01",  "Land O Lakes",                         "Dairy products supplier",                      "4001 Lexington Ave N",  None,        None, "Arden Hills",      "MN", "55126", "US", "001-2100"],
    ["SCHEP01",  "Schepps Dairy",                        "Regional dairy distributor",                   "100 Schepps Blvd",      None,        None, "Dallas",           "TX", "75229", "US", "001-2100"],
    ["KEHNE01",  "Kehe Distributors",                    "Natural and specialty grocery",                "900 S Milwaukee Ave",   None,        None, "Wheeling",         "IL", "60090", "US", "001-2100"],
    ["EMPBK01",  "Empire Bakery",                        "Artisan bread and pastry supplier",            "521 5th Ave",           "Ste 1800",  None, "New York",         "NY", "10175", "US", "001-2100"],
    ["OTTFD01",  "Otis Spunkmeyer",                      "Bakery and dessert products",                  "7090 Mowry Ave",        None,        None, "Newark",           "CA", "94560", "US", "001-2100"],
    ["REPNB01",  "Republic National Distributing",       "Wine and spirits distributor",                 "5000 Industry Park Dr", None,        None, "Smyrna",           "GA", "30080", "US", "002-2100"],
    ["GLAZR01",  "Glazers Distributors",                 "Beer wine and spirits",                        "14911 Quorum Dr",       "Ste 300",   None, "Dallas",           "TX", "75254", "US", "002-2100"],
    ["SOGLZ01",  "Southern Glazers Wine and Spirits",    "National beverage distributor",                "3501 NW 107th Ave",     None,        None, "Doral",            "FL", "33178", "US", "002-2100"],
    ["REYES01",  "Reyes Beer Division",                  "Beer and malt beverage distributor",           "6250 N River Rd",       None,        None, "Rosemont",         "IL", "60018", "US", "002-2100"],
    ["CCSW01",   "Coca-Cola Southwest Beverages",        "Soft drink and fountain supplier",             "2025 McKelvey Rd",      None,        None, "Maryland Heights",  "MO", "63043", "US", "001-2100"],
    ["PEPSI01",  "PepsiCo Foodservice",                  "Beverages and snack foods",                    "700 Anderson Hill Rd",  None,        None, "Purchase",         "NY", "10577", "US", "001-2100"],
    ["ECLAB01",  "Ecolab Inc",                           "Cleaning and sanitation products",             "1 Ecolab Pl",           None,        None, "Saint Paul",       "MN", "55102", "US", "001-2100"],
    ["CINTA01",  "Cintas Corporation",                   "Uniforms and linen services",                  "6800 Cintas Blvd",      None,        None, "Mason",            "OH", "45040", "US", "001-2100"],
    ["ARMKU01",  "Aramark Uniform Services",             "Uniform rental and laundry",                   "1101 Market St",        None,        None, "Philadelphia",     "PA", "19107", "US", "001-2100"],
    ["NTRST01",  "National Restaurant Supply",           "Commercial kitchen equipment",                 "3850 Forest Ln",        None,        None, "Dallas",           "TX", "75234", "US", "001-2100"],
    ["WEBST01",  "WebstaurantStore",                     "Restaurant equipment and supplies",            "40 24th St",            None,        None, "Pittston",         "PA", "18640", "US", "001-2100"],
    ["VOLLR01",  "Vollrath Company",                     "Commercial cooking equipment",                 "1236 N 18th St",        None,        None, "Sheboygan",        "WI", "53081", "US", "001-2100"],
    ["CAMBO01",  "Cambro Manufacturing",                 "Foodservice storage and transport",            "5801 Skylab Rd",        None,        None, "Huntington Beach", "CA", "92647", "US", "001-2100"],
    ["WELBT01",  "Welbilt Inc",                          "Commercial foodservice equipment",             "2227 Welbilt Blvd",     None,        None, "New Port Richey",  "FL", "34655", "US", "001-2100"],
    ["RSRTE01",  "Restaurant Technologies",              "Cooking oil management systems",               "2250 Pilot Knob Rd",    None,        None, "Mendota Heights",  "MN", "55120", "US", "001-2100"],
    ["NCRCR01",  "NCR Corporation",                      "POS systems and technology",                   "864 Spring St NW",      None,        None, "Atlanta",          "GA", "30308", "US", "001-2100"],
    ["TOAST01",  "Toast Inc",                            "Restaurant POS and management software",       "401 Park Dr",           "Ste 801",   None, "Boston",           "MA", "02215", "US", "001-2100"],
    ["HRTLD01",  "Heartland Payment Systems",            "Payment processing services",                  "90 Nassau St",          None,        None, "Princeton",        "NJ", "08542", "US", "001-2100"],
    ["AMEXC01",  "American Express",                     "Corporate credit card services",               "200 Vesey St",          None,        None, "New York",         "NY", "10285", "US", "001-2100"],
    ["OPNTR01",  "OpenTable Inc",                        "Reservation management platform",              "1 Montgomery St",       "Ste 700",   None, "San Francisco",    "CA", "94104", "US", "001-2100"],
    ["INTUT01",  "Intuit QuickBooks",                    "Accounting software subscription",             "2700 Coast Ave",        None,        None, "Mountain View",    "CA", "94043", "US", "001-2100"],
    ["RPSVC01",  "Republic Services",                    "Waste management and recycling",               "18500 N Allied Way",    None,        None, "Phoenix",          "AZ", "85054", "US", "001-2100"],
    ["ROLLI01",  "Rollins Pest Control",                 "Commercial pest control services",             "2170 Piedmont Rd NE",   None,        None, "Atlanta",          "GA", "30324", "US", "001-2100"],
    ["HRTFD01",  "Hartford Financial Services",          "Commercial insurance",                         "1 Hartford Plz",        None,        None, "Hartford",         "CT", "06155", "US", "001-2100"],
    ["LOCKTN01", "Lockton Companies",                    "Restaurant and hospitality insurance",         "444 W 47th St",         "Ste 900",   None, "Kansas City",      "MO", "64112", "US", "001-2100"],
    ["SPLXG01",  "Simplex-Grinnell",                     "Fire suppression and safety",                  "50 Technology Dr",      None,        None, "Westminster",      "MA", "01441", "US", "001-2100"],
    ["VECTS01",  "Vector Security",                      "Security and surveillance systems",            "2000 Ericsson Dr",      None,        None, "Warrendale",       "PA", "15086", "US", "001-2100"],
    ["ZAYOG01",  "Zayo Group",                           "Fiber internet and connectivity",              "1821 30th St",          None,        None, "Boulder",          "CO", "80301", "US", "001-2100"],
    ["ATLGL01",  "Atlanta Gas Light",                    "Natural gas utility",                          "10 Peachtree Pl NE",    None,        None, "Atlanta",          "GA", "30309", "US", "001-2100"],
    ["ONCOR01",  "Oncor Electric Delivery",              "Electric utility services",                    "1616 Woodall Rodgers Fwy", None,     None, "Dallas",           "TX", "75202", "US", "001-2100"],
    ["EVRGI01",  "Evergreen Landscape Services",         "Commercial landscaping",                       "9320 E Raintree Dr",    "Ste 100",   None, "Scottsdale",       "AZ", "85260", "US", "001-2100"],
    ["SEALE01",  "Sealand Food",                         "Specialty seafood distributor",                "3420 Ocean Dr",         None,        None, "Virginia Beach",   "VA", "23451", "US", "001-2100"],
    ["SIMEU01",  "Simeus Foods",                         "Specialty protein and seafood",                "400 E Las Colinas Blvd","Ste 200",   None, "Irving",           "TX", "75039", "US", "001-2100"],
    ["CHURF01",  "Chur Farms Produce",                   "Local fresh produce supplier",                 "4500 Farm Rd 55",       None,        None, "Hereford",         "TX", "79045", "US", "001-2100"],
]
headers_v = ["VendorCode","Name","Description","Address1","Address2","Address3","City","State","PostalCode","Country","DefaultPayablesAccount"]
style_sheet(ws, headers_v, [[c if c is not None else "" for c in r] for r in vendors])
wb.save(os.path.join(DIR, "Vendors.xlsx"))
print("Vendors.xlsx saved")

# Remove old csv files
for f in ["EodSections.csv", "EodColumns.csv", "EodRows.csv", "ChartOfAccounts.csv"]:
    p = os.path.join(DIR, f)
    if os.path.exists(p):
        os.remove(p)
        print(f"Removed {f}")

print("Done.")
