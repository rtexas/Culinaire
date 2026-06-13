from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR_BG = "2B6B35"; HDR_FG = "FFFFFF"; ROW_ALT = "F0F7F0"; FN = "Arial"

def hdr(cell, h):
    cell.value = h
    cell.font = Font(name=FN, bold=True, color=HDR_FG, size=10)
    cell.fill = PatternFill("solid", start_color=HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def body(ws, r, row, n):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name=FN, size=10)
        cell.alignment = Alignment(vertical="center")
    if r % 2 == 0:
        for c in range(1, n+1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", start_color=ROW_ALT)

def autofit(ws, hdrs, rows):
    for i, h in enumerate(hdrs, 1):
        ml = len(str(h))
        for row in rows:
            ml = max(ml, len(str(row[i-1]) if i-1 < len(row) else ""))
        ws.column_dimensions[get_column_letter(i)].width = min(ml+4, 50)

# ── CHART OF ACCOUNTS ────────────────────────────────────────────────────────
coa_rows = [
    ("01-000-1000","Cash - Operating Account","Assets","Balance Sheet"),
    ("01-000-1010","Cash - Payroll Account","Assets","Balance Sheet"),
    ("01-000-1020","Petty Cash","Assets","Balance Sheet"),
    ("01-000-1100","Accounts Receivable","Assets","Balance Sheet"),
    ("01-000-1110","Credit Card Receivable","Assets","Balance Sheet"),
    ("01-000-1200","Food Inventory","Assets","Balance Sheet"),
    ("01-000-1210","Beverage Inventory","Assets","Balance Sheet"),
    ("01-000-1220","Bar Supplies Inventory","Assets","Balance Sheet"),
    ("01-000-1300","Prepaid Insurance","Assets","Balance Sheet"),
    ("01-000-1310","Prepaid Rent","Assets","Balance Sheet"),
    ("01-000-1400","Furniture and Fixtures","Assets","Balance Sheet"),
    ("01-000-1410","Kitchen Equipment","Assets","Balance Sheet"),
    ("01-000-1420","Leasehold Improvements","Assets","Balance Sheet"),
    ("01-000-1490","Accumulated Depreciation","Assets","Balance Sheet"),
    ("01-000-2000","Accounts Payable","Liabilities","Balance Sheet"),
    ("01-000-2010","Accrued Payroll","Liabilities","Balance Sheet"),
    ("01-000-2020","Accrued Payroll Taxes","Liabilities","Balance Sheet"),
    ("01-000-2030","Sales Tax Payable","Liabilities","Balance Sheet"),
    ("01-000-2040","Gratuity Payable","Liabilities","Balance Sheet"),
    ("01-000-2050","Gift Card Liability","Liabilities","Balance Sheet"),
    ("01-000-2100","Notes Payable Short Term","Liabilities","Balance Sheet"),
    ("01-000-2200","Notes Payable Long Term","Liabilities","Balance Sheet"),
    ("01-000-3000","Owner Equity","Equity","Balance Sheet"),
    ("01-000-3100","Retained Earnings","Equity","Balance Sheet"),
    ("01-000-3200","Current Year Earnings","Equity","Balance Sheet"),
    ("02-200-4000","Food Sales Main Dining","Revenue","Income Statement"),
    ("02-200-4010","Beverage Sales Main Dining","Revenue","Income Statement"),
    ("02-200-4020","Private Event Revenue Main Dining","Revenue","Income Statement"),
    ("02-200-4030","Catering Revenue Main Dining","Revenue","Income Statement"),
    ("02-200-4900","Other Revenue Main Dining","Revenue","Income Statement"),
    ("03-300-4000","Food Sales Bar and Lounge","Revenue","Income Statement"),
    ("03-300-4010","Beverage Sales Bar and Lounge","Revenue","Income Statement"),
    ("03-300-4020","Entertainment Revenue Bar and Lounge","Revenue","Income Statement"),
    ("03-300-4900","Other Revenue Bar and Lounge","Revenue","Income Statement"),
    ("02-100-5000","Food Cost Kitchen","Cost of Sales","Income Statement"),
    ("02-100-5010","Beverage Cost Kitchen","Cost of Sales","Income Statement"),
    ("02-100-5020","Paper and Supplies Cost Kitchen","Cost of Sales","Income Statement"),
    ("02-200-5030","Comps and Voids Main Dining","Cost of Sales","Income Statement"),
    ("03-300-5000","Food Cost Bar","Cost of Sales","Income Statement"),
    ("03-300-5010","Beverage Cost Bar","Cost of Sales","Income Statement"),
    ("03-300-5020","Bar Supplies Cost","Cost of Sales","Income Statement"),
    ("03-300-5030","Comps and Voids Bar","Cost of Sales","Income Statement"),
    ("01-400-6000","Management Salary Corporate","Labor","Income Statement"),
    ("01-400-6010","Administrative Salary","Labor","Income Statement"),
    ("01-400-6100","Payroll Taxes Corporate","Labor","Income Statement"),
    ("01-400-6110","Employee Benefits Corporate","Labor","Income Statement"),
    ("01-400-6120","Workers Compensation Corporate","Labor","Income Statement"),
    ("02-100-6000","Kitchen Labor Hourly","Labor","Income Statement"),
    ("02-100-6010","Kitchen Labor Overtime","Labor","Income Statement"),
    ("02-200-6000","Front of House Labor Hourly","Labor","Income Statement"),
    ("02-200-6010","Front of House Labor Overtime","Labor","Income Statement"),
    ("02-000-6020","Management Salary Main Dining","Labor","Income Statement"),
    ("02-000-6100","Payroll Taxes Main Dining","Labor","Income Statement"),
    ("02-000-6110","Employee Benefits Main Dining","Labor","Income Statement"),
    ("02-000-6120","Workers Compensation Main Dining","Labor","Income Statement"),
    ("03-300-6000","Bar Labor Hourly","Labor","Income Statement"),
    ("03-300-6010","Bar Labor Overtime","Labor","Income Statement"),
    ("03-000-6020","Management Salary Bar and Lounge","Labor","Income Statement"),
    ("03-000-6100","Payroll Taxes Bar and Lounge","Labor","Income Statement"),
    ("03-000-6110","Employee Benefits Bar and Lounge","Labor","Income Statement"),
    ("01-500-7000","Rent and Lease Corporate","Operating Expenses","Income Statement"),
    ("01-500-7010","Utilities Corporate","Operating Expenses","Income Statement"),
    ("01-500-7020","Telephone and Internet","Operating Expenses","Income Statement"),
    ("01-400-7100","Marketing and Advertising","Operating Expenses","Income Statement"),
    ("01-400-7110","Social Media and Digital Marketing","Operating Expenses","Income Statement"),
    ("01-500-7200","Repairs and Maintenance","Operating Expenses","Income Statement"),
    ("01-400-7300","Office Supplies","Operating Expenses","Income Statement"),
    ("01-400-7400","Insurance General Liability","Operating Expenses","Income Statement"),
    ("01-400-7410","Insurance Property","Operating Expenses","Income Statement"),
    ("01-400-7500","Credit Card Processing Fees","Operating Expenses","Income Statement"),
    ("01-400-7600","Accounting and Legal Fees","Operating Expenses","Income Statement"),
    ("01-400-7700","Licenses and Permits","Operating Expenses","Income Statement"),
    ("01-400-7800","Training and Education","Operating Expenses","Income Statement"),
    ("02-200-7300","FOH Supplies Main Dining","Operating Expenses","Income Statement"),
    ("02-100-7310","Kitchen Supplies Main Dining","Operating Expenses","Income Statement"),
    ("02-200-7600","Uniforms Main Dining","Operating Expenses","Income Statement"),
    ("02-500-7200","Repairs and Maintenance Main Dining","Operating Expenses","Income Statement"),
    ("02-500-7900","Pest Control Main Dining","Operating Expenses","Income Statement"),
    ("03-300-7300","Bar Supplies Consumables","Operating Expenses","Income Statement"),
    ("03-300-7600","Uniforms Bar and Lounge","Operating Expenses","Income Statement"),
    ("03-500-7200","Repairs and Maintenance Bar","Operating Expenses","Income Statement"),
    ("03-500-7900","Pest Control Bar","Operating Expenses","Income Statement"),
    ("01-000-8000","Interest Income","Other Income and Expense","Income Statement"),
    ("01-000-8100","Interest Expense","Other Income and Expense","Income Statement"),
    ("01-000-8200","Gain Loss on Disposal of Assets","Other Income and Expense","Income Statement"),
    ("01-000-8900","Miscellaneous Income","Other Income and Expense","Income Statement"),
]

coa_hdrs = ["Account Name","Account Description","Account Category","Account Type"]
wb = Workbook(); ws = wb.active; ws.title = "Chart of Accounts"
ws.row_dimensions[1].height = 20
for c, h in enumerate(coa_hdrs, 1): hdr(ws.cell(1, c), h)
for r, row in enumerate(coa_rows, 2): body(ws, r, row, 4)
autofit(ws, coa_hdrs, coa_rows)
ws.freeze_panes = "A2"
wb.save(r"C:\ClaudeOutput\Culinaire\SampleData\Chart of Accounts.xlsx")
print(f"COA saved: {len(coa_rows)} rows")

# ── VENDORS ──────────────────────────────────────────────────────────────────
ven_rows = [
    ("SYSCO01","Sysco Corporation","Broadline food distributor","1390 Enclave Pkwy","","","Houston","TX","77077","US","01-000-2000"),
    ("USFDS01","US Foods","National broadline distributor","9399 W Higgins Rd","","","Rosemont","IL","60018","US","01-000-2000"),
    ("PFGFT01","Performance Food Group","Broadline food and beverage distributor","12500 W Creek Pkwy","","","Richmond","VA","23238","US","01-000-2000"),
    ("GPFSC01","Gordon Food Service","Broadline food distributor","1300 Gezon Pkwy SW","","","Wyoming","MI","49509","US","01-000-2000"),
    ("REJFS01","Reinhart Foodservice","Regional broadline distributor","6250 N Wilson Dr","","","Milwaukee","WI","53223","US","01-000-2000"),
    ("BENCO01","Ben E Keith Foods","Southwest food distributor","601 E 7th St","","","Fort Worth","TX","76102","US","01-000-2000"),
    ("DOLE01","Dole Food Company","Fresh produce supplier","1000 Corporate Ctr Dr","","","Monterey Park","CA","91754","US","01-000-2000"),
    ("TAYFM01","Taylor Farms","Fresh-cut produce and salads","10 Becerra Rd","","","Salinas","CA","93905","US","01-000-2000"),
    ("CARGM01","Cargill Meat Solutions","Beef and protein supplier","151 N Main St","","","Wichita","KS","67202","US","01-000-2000"),
    ("TYSON01","Tyson Foods","Poultry and protein distributor","2200 Don Tyson Pkwy","","","Springdale","AR","72762","US","01-000-2000"),
    ("NBRKE01","Nebraska Beef","Premium beef supplier","605 S 3rd St","","","Omaha","NE","68102","US","01-000-2000"),
    ("LANDB01","Land O Lakes","Dairy products supplier","4001 Lexington Ave N","","","Arden Hills","MN","55126","US","01-000-2000"),
    ("SCHEP01","Schepps Dairy","Regional dairy distributor","100 Schepps Blvd","","","Dallas","TX","75229","US","01-000-2000"),
    ("KEHNE01","Kehe Distributors","Natural and specialty grocery","900 S Milwaukee Ave","","","Wheeling","IL","60090","US","01-000-2000"),
    ("EMPBK01","Empire Bakery","Artisan bread and pastry supplier","521 5th Ave","Ste 1800","","New York","NY","10175","US","01-000-2000"),
    ("OTTFD01","Otis Spunkmeyer","Bakery and dessert products","7090 Mowry Ave","","","Newark","CA","94560","US","01-000-2000"),
    ("REPNB01","Republic National Distributing","Wine and spirits distributor","5000 Industry Park Dr","","","Smyrna","GA","30080","US","01-000-2000"),
    ("GLAZR01","Glazers Distributors","Beer wine and spirits","14911 Quorum Dr","Ste 300","","Dallas","TX","75254","US","01-000-2000"),
    ("SOGLZ01","Southern Glazers Wine and Spirits","National beverage distributor","3501 NW 107th Ave","","","Doral","FL","33178","US","01-000-2000"),
    ("REYES01","Reyes Beer Division","Beer and malt beverage distributor","6250 N River Rd","","","Rosemont","IL","60018","US","01-000-2000"),
    ("CCSW01","Coca-Cola Southwest Beverages","Soft drink and fountain supplier","2025 McKelvey Rd","","","Maryland Heights","MO","63043","US","01-000-2000"),
    ("PEPSI01","PepsiCo Foodservice","Beverages and snack foods","700 Anderson Hill Rd","","","Purchase","NY","10577","US","01-000-2000"),
    ("ECLAB01","Ecolab Inc","Cleaning and sanitation products","1 Ecolab Pl","","","Saint Paul","MN","55102","US","01-000-2000"),
    ("CINTA01","Cintas Corporation","Uniforms and linen services","6800 Cintas Blvd","","","Mason","OH","45040","US","01-000-2000"),
    ("ARMKU01","Aramark Uniform Services","Uniform rental and laundry","1101 Market St","","","Philadelphia","PA","19107","US","01-000-2000"),
    ("NTRST01","National Restaurant Supply","Commercial kitchen equipment","3850 Forest Ln","","","Dallas","TX","75234","US","01-000-2000"),
    ("WEBST01","WebstaurantStore","Restaurant equipment and supplies","40 24th St","","","Pittston","PA","18640","US","01-000-2000"),
    ("VOLLR01","Vollrath Company","Commercial cooking equipment","1236 N 18th St","","","Sheboygan","WI","53081","US","01-000-2000"),
    ("CAMBO01","Cambro Manufacturing","Foodservice storage and transport","5801 Skylab Rd","","","Huntington Beach","CA","92647","US","01-000-2000"),
    ("WELBT01","Welbilt Inc","Commercial foodservice equipment","2227 Welbilt Blvd","","","New Port Richey","FL","34655","US","01-000-2000"),
    ("RSRTE01","Restaurant Technologies","Cooking oil management systems","2250 Pilot Knob Rd","","","Mendota Heights","MN","55120","US","01-000-2000"),
    ("NCRCR01","NCR Corporation","POS systems and technology","864 Spring St NW","","","Atlanta","GA","30308","US","01-000-2000"),
    ("TOAST01","Toast Inc","Restaurant POS and management software","401 Park Dr","Ste 801","","Boston","MA","02215","US","01-000-2000"),
    ("HRTLD01","Heartland Payment Systems","Payment processing services","90 Nassau St","","","Princeton","NJ","08542","US","01-000-2000"),
    ("AMEXC01","American Express","Corporate credit card services","200 Vesey St","","","New York","NY","10285","US","01-000-2000"),
    ("OPNTR01","OpenTable Inc","Reservation management platform","1 Montgomery St","Ste 700","","San Francisco","CA","94104","US","01-000-2000"),
    ("INTUT01","Intuit QuickBooks","Accounting software subscription","2700 Coast Ave","","","Mountain View","CA","94043","US","01-000-2000"),
    ("RPSVC01","Republic Services","Waste management and recycling","18500 N Allied Way","","","Phoenix","AZ","85054","US","01-000-2000"),
    ("ROLLI01","Rollins Pest Control","Commercial pest control services","2170 Piedmont Rd NE","","","Atlanta","GA","30324","US","01-000-2000"),
    ("HRTFD01","Hartford Financial Services","Commercial insurance","1 Hartford Plz","","","Hartford","CT","06155","US","01-000-2000"),
    ("LOCKTN01","Lockton Companies","Restaurant and hospitality insurance","444 W 47th St","Ste 900","","Kansas City","MO","64112","US","01-000-2000"),
    ("SPLXG01","Simplex-Grinnell","Fire suppression and safety","50 Technology Dr","","","Westminster","MA","01441","US","01-000-2000"),
    ("VECTS01","Vector Security","Security and surveillance systems","2000 Ericsson Dr","","","Warrendale","PA","15086","US","01-000-2000"),
    ("ZAYOG01","Zayo Group","Fiber internet and connectivity","1821 30th St","","","Boulder","CO","80301","US","01-000-2000"),
    ("ATLGL01","Atlanta Gas Light","Natural gas utility","10 Peachtree Pl NE","","","Atlanta","GA","30309","US","01-000-2000"),
    ("ONCOR01","Oncor Electric Delivery","Electric utility services","1616 Woodall Rodgers Fwy","","","Dallas","TX","75202","US","01-000-2000"),
    ("EVRGI01","Evergreen Landscape Services","Commercial landscaping","9320 E Raintree Dr","Ste 100","","Scottsdale","AZ","85260","US","01-000-2000"),
    ("SEALE01","Sealand Food","Specialty seafood distributor","3420 Ocean Dr","","","Virginia Beach","VA","23451","US","01-000-2000"),
    ("SIMEU01","Simeus Foods","Specialty protein and seafood","400 E Las Colinas Blvd","Ste 200","","Irving","TX","75039","US","01-000-2000"),
    ("CHURF01","Chur Farms Produce","Local fresh produce supplier","4500 Farm Rd 55","","","Hereford","TX","79045","US","01-000-2000"),
]

ven_hdrs = ["VendorCode","Name","Description","Address1","Address2","Address3","City","State","PostalCode","Country","DefaultPayablesAccount"]
wb2 = Workbook(); ws2 = wb2.active; ws2.title = "Vendors"
ws2.row_dimensions[1].height = 20
for c, h in enumerate(ven_hdrs, 1): hdr(ws2.cell(1, c), h)
for r, row in enumerate(ven_rows, 2): body(ws2, r, row, 11)
autofit(ws2, ven_hdrs, ven_rows)
ws2.freeze_panes = "A2"
wb2.save(r"C:\ClaudeOutput\Culinaire\SampleData\Vendors.xlsx")
print(f"Vendors saved: {len(ven_rows)} rows")

# ── COA SEGMENTS ─────────────────────────────────────────────────────────────
# Account format: XX-XXX-XXXX  (Entity - Location - Account Number)
seg_rows = [
    (1, "Entity"),
    (2, "Location"),
    (3, "Account Number"),
]

seg_hdrs = ["Segment Number", "Description"]
wb3 = Workbook(); ws3 = wb3.active; ws3.title = "CoA Segments"
ws3.row_dimensions[1].height = 20
for c, h in enumerate(seg_hdrs, 1): hdr(ws3.cell(1, c), h)
for r, row in enumerate(seg_rows, 2): body(ws3, r, row, 2)
autofit(ws3, seg_hdrs, seg_rows)
ws3.freeze_panes = "A2"
wb3.save(r"C:\ClaudeOutput\Culinaire\SampleData\CoA Segments.xlsx")
print(f"CoA Segments saved: {len(seg_rows)} rows")
