from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR_BG = "2B6B35"; HDR_FG = "FFFFFF"; FN = "Arial"

def hdr(cell, h):
    cell.value = h
    cell.font = Font(name=FN, bold=True, color=HDR_FG, size=10)
    cell.fill = PatternFill("solid", start_color=HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")

hdrs = ["Name", "Code", "Description"]
rows = [("United States", "US", "United States of America")]

wb = Workbook(); ws = wb.active; ws.title = "Countries"
ws.row_dimensions[1].height = 20

for c, h in enumerate(hdrs, 1): hdr(ws.cell(1, c), h)

for c, v in enumerate(rows[0], 1):
    cell = ws.cell(row=2, column=c, value=v)
    cell.font = Font(name=FN, size=10)
    cell.alignment = Alignment(vertical="center")

for i, h in enumerate(hdrs, 1):
    ml = max(len(h), len(str(rows[0][i-1])))
    ws.column_dimensions[get_column_letter(i)].width = ml + 4

ws.freeze_panes = "A2"

path = r"C:\ClaudeOutput\Culinaire\SampleData\Countries.xlsx"
wb.save(path)
print(f"Saved: {path}")
