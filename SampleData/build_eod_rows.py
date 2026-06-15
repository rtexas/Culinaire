import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# (Name, Section, Description)
rows = [
    # ── Sales (+1) — food & beverage revenue ──────────────────────────────
    ("Food Sales",               "Sales", "Total food revenue for the period"),
    ("Beverage Sales",           "Sales", "Total non-alcoholic beverage revenue"),
    ("Beer Sales",               "Sales", "Draft and bottled beer revenue"),
    ("Wine Sales",               "Sales", "Bottled and by-the-glass wine revenue"),
    ("Liquor Sales",             "Sales", "Spirits and cocktail revenue"),
    ("Appetizers",               "Sales", "Starter and appetizer category sales"),
    ("Entrees",                  "Sales", "Main course / entrée category sales"),
    ("Desserts",                 "Sales", "Dessert category sales"),
    ("Non-Alcoholic Beverages",  "Sales", "Soft drinks, juices, coffee, and tea revenue"),
    ("Catering Sales",           "Sales", "Off-site and on-site catering revenue"),
    ("Private Event Sales",      "Sales", "Private dining and buyout event revenue"),
    ("Delivery Sales",           "Sales", "Third-party and in-house delivery revenue"),
    ("Takeout Sales",            "Sales", "To-go and curbside pickup revenue"),
    ("Gift Card Sales",          "Sales", "Gift card purchases for the period"),
    ("Merchandise Sales",        "Sales", "Branded merchandise and retail sales"),

    # ── Dispositions (-1) — tender types & reductions ─────────────────────
    ("Cash",                       "Dispositions", "Cash payments collected for the period"),
    ("Visa",                       "Dispositions", "Visa credit and debit card payments collected"),
    ("MasterCard",                 "Dispositions", "MasterCard credit and debit card payments collected"),
    ("Discover",                   "Dispositions", "Discover card payments collected"),
    ("Amex",                       "Dispositions", "American Express card payments collected"),
    ("Checks",                     "Dispositions", "Personal and business check payments collected"),
    ("Comps",                      "Dispositions", "Complimentary items and checks written off"),
    ("Voids",                      "Dispositions", "Voided transactions removed from revenue"),
    ("Discounts",                  "Dispositions", "Promotional and loyalty discounts applied"),
    ("Employee Meals",             "Dispositions", "Employee meal credits and staff dining deductions"),
    ("Waste",                      "Dispositions", "Food and beverage waste written off"),

    # ── Liabilities (0) — expenses & obligations ───────────────────────────
    ("Sales Tax",                "Liabilities", "Sales tax collected and due to state/local authority"),
    ("Gratuity",                 "Liabilities", "Auto-gratuity and service charge collected"),
    ("Landscape",                "Liabilities", "Grounds maintenance and landscaping expense"),
    ("Building",                 "Liabilities", "Building rent, lease, or mortgage expense"),
    ("Equipment",                "Liabilities", "Kitchen and facility equipment expense"),
    ("Utilities",                "Liabilities", "Electric, gas, and water expense"),
    ("Insurance",                "Liabilities", "General liability and property insurance expense"),
    ("Repairs & Maintenance",    "Liabilities", "Facility and equipment repair expense"),
    ("Supplies",                 "Liabilities", "Operating supply and consumable expense"),
    ("Credit Card Fees",         "Liabilities", "Merchant processing and transaction fees"),

    # ── Unit Accounts (0/excluded) — statistical counts ────────────────────
    ("Count of Customers",       "Unit Accounts", "Total covers / guests served for the period"),
    ("Count of Checks",          "Unit Accounts", "Total number of guest checks or tickets"),
    ("Count of Tables Turned",   "Unit Accounts", "Total table turns for the period"),
    ("Count of Takeout Orders",  "Unit Accounts", "Total to-go and curbside orders"),
    ("Count of Delivery Orders", "Unit Accounts", "Total third-party and in-house delivery orders"),
    ("Count of Bar Covers",      "Unit Accounts", "Total bar-seating covers for the period"),
    ("Labor Hours Kitchen",      "Unit Accounts", "Total kitchen labor hours clocked for the period"),
    ("Labor Hours Front of House","Unit Accounts","Total FOH labor hours clocked for the period"),
    ("Labor Hours Bar",          "Unit Accounts", "Total bar labor hours clocked for the period"),
    ("Count of Voids",           "Unit Accounts", "Total number of voided items or checks"),
    ("Count of Comps",           "Unit Accounts", "Total number of complimentary items or checks"),
    ("Count of Reservations",    "Unit Accounts", "Total reservations booked for the period"),
    ("Count of No Shows",        "Unit Accounts", "Total reservation no-shows for the period"),
    ("Count of Walk Ins",        "Unit Accounts", "Total walk-in parties seated"),
    ("Count of Catering Events", "Unit Accounts", "Total catering events executed for the period"),
    ("Count of Private Dining Events","Unit Accounts","Total private dining buyout events"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EOD Rows"

HDR_BG = "2B6B35"; HDR_FG = "FFFFFF"; ALT = "F0F7F0"
header_fill  = PatternFill("solid", fgColor=HDR_BG)
header_font  = Font(name="Arial", bold=True, color=HDR_FG, size=11)
header_align = Alignment(horizontal="center", vertical="center")
data_font    = Font(name="Arial", size=10)
thin         = Border(
    left=Side(style="thin", color="C8DFC8"), right=Side(style="thin", color="C8DFC8"),
    top=Side(style="thin", color="C8DFC8"),  bottom=Side(style="thin", color="C8DFC8"),
)

# Section colour bands
section_fills = {
    "Sales":         PatternFill("solid", fgColor="EAF4EA"),
    "Income":        PatternFill("solid", fgColor="EAF0FA"),
    "Dispositions":  PatternFill("solid", fgColor="FFF0EA"),
    "Liabilities":   PatternFill("solid", fgColor="FAF4EA"),
    "Unit Accounts": PatternFill("solid", fgColor="F4EAF4"),
}
section_fonts = {
    "Sales":         Font(name="Arial", size=10, color="1A5C1A"),
    "Income":        Font(name="Arial", size=10, color="1A3A6B"),
    "Dispositions":  Font(name="Arial", size=10, color="7A3010"),
    "Liabilities":   Font(name="Arial", size=10, color="6B4A1A"),
    "Unit Accounts": Font(name="Arial", size=10, color="5C1A5C"),
}

headers    = ["Name", "Section", "Description"]
col_widths = [28,     16,        55]

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = header_align; cell.border = thin
    ws.column_dimensions[cell.column_letter].width = w
ws.row_dimensions[1].height = 20

for row_idx, (name, section, desc) in enumerate(rows, start=2):
    row_fill = section_fills.get(section, PatternFill())
    row_font = section_fonts.get(section, Font(name="Arial", size=10))
    for col, val in enumerate([name, section, desc], start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = row_font; cell.fill = row_fill; cell.border = thin
        cell.alignment = Alignment(vertical="center")

ws.freeze_panes = "A2"
out = r"C:\ClaudeOutput\Culinaire\SampleData\EOD Rows.xlsx"
wb.save(out)
print(f"Saved: {out}  ({len(rows)} rows)")
print(f"  Sales:         {sum(1 for _,s,_ in rows if s=='Sales')}")
print(f"  Dispositions:  {sum(1 for _,s,_ in rows if s=='Dispositions')}")
print(f"  Liabilities:   {sum(1 for _,s,_ in rows if s=='Liabilities')}")
print(f"  Unit Accounts: {sum(1 for _,s,_ in rows if s=='Unit Accounts')}")
