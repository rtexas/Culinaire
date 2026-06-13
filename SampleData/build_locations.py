import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# CoA uses segment 2 for department/location codes (3-digit numeric).
# Segment values already in use: 000=Corp, 100=Kitchen, 200=Main Dining,
# 300=Bar/Lounge, 400=Corporate/Admin, 500=Catering.
# Additional codes 110-590 used for sub-locations.

locations = [
    # code,  name,                         description,                                   seg
    ("000",  "Corporate Office",           "Corporate and consolidated entity",             2),
    ("100",  "Kitchen - Main",             "Primary production kitchen",                    2),
    ("110",  "Kitchen - Prep",             "Cold prep and pastry production",               2),
    ("115",  "Kitchen - Banquet",          "Banquet and event production kitchen",          2),
    ("200",  "Main Dining Room",           "Full-service main dining room",                 2),
    ("210",  "Private Dining Room A",      "Small private dining room, seats 12",           2),
    ("215",  "Private Dining Room B",      "Large private dining room, seats 30",           2),
    ("220",  "Patio and Terrace",          "Outdoor patio dining and terrace seating",      2),
    ("300",  "Bar and Lounge",             "Full-service bar and cocktail lounge",          2),
    ("310",  "Rooftop Bar",               "Seasonal rooftop cocktail venue",               2),
    ("315",  "Sports Bar",                "High-volume sports bar with AV screens",        2),
    ("400",  "Corporate Administration",  "Corporate management and admin department",     2),
    ("410",  "Human Resources",           "HR and payroll administration",                 2),
    ("415",  "Finance and Accounting",    "Accounting, AP, and financial reporting",       2),
    ("420",  "Marketing",                 "Marketing, social media, and promotions",       2),
    ("425",  "IT and Systems",            "Technology infrastructure and systems",         2),
    ("500",  "Catering Operations",       "Off-site catering and event delivery hub",      2),
    ("510",  "Hotel Grand - Dining",      "Main dining room at Hotel Grand",               2),
    ("515",  "Hotel Grand - Banquet",     "Banquet and event catering at Hotel Grand",     2),
    ("520",  "Convention Center",         "High-volume catering for conventions",          2),
    ("530",  "Airport Terminal A",        "Full-service dining at Terminal A departures",  2),
    ("535",  "Airport Terminal B",        "Quick-service counter at Terminal B gates",     2),
    ("540",  "University Campus",         "Campus dining for students and faculty",        2),
    ("550",  "Medical Center Cafeteria",  "Cafeteria and patient meal service",            2),
    ("590",  "Pop-Up and Mobile",         "Mobile and temporary event kitchen",            2),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Locations"

header_fill  = PatternFill("solid", fgColor="2B6B35")
header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center")
alt_fill     = PatternFill("solid", fgColor="F0F7F0")
data_font    = Font(name="Arial", size=10)
thin_border  = Border(
    left   = Side(style="thin", color="C8DFC8"),
    right  = Side(style="thin", color="C8DFC8"),
    top    = Side(style="thin", color="C8DFC8"),
    bottom = Side(style="thin", color="C8DFC8"),
)

headers    = ["Code", "Name", "Description", "Segment Number"]
col_widths = [10,     38,     55,             16]

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align
    cell.border    = thin_border
    ws.column_dimensions[cell.column_letter].width = w

ws.row_dimensions[1].height = 20

for row_idx, (code, name, desc, seg) in enumerate(locations, start=2):
    fill = alt_fill if row_idx % 2 == 0 else PatternFill()
    for col, val in enumerate([code, name, desc, seg], start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font      = data_font
        cell.fill      = fill
        cell.border    = thin_border
        cell.alignment = Alignment(vertical="center",
                                   horizontal="center" if col in (1, 4) else "left")

ws.freeze_panes = "A2"

out = r"C:\ClaudeOutput\Culinaire\SampleData\Locations.xlsx"
wb.save(out)
print(f"Saved: {out}  ({len(locations)} locations)")
