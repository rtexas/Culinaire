from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL  = PatternFill("solid", fgColor="2B6B35")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="F0F7F0")
NORMAL_FONT  = Font(name="Arial", size=10)
THIN         = Side(style="thin", color="CCCCCC")
BORDER       = Border(bottom=THIN)

items = [
    ("BURG01", "Classic Cheeseburger",          "8 oz beef patty with American cheese, lettuce, tomato, onion",    12.99),
    ("BURG02", "Bacon Mushroom Burger",          "8 oz beef patty topped with applewood bacon and sautéed mushrooms", 14.99),
    ("CHIX01", "Grilled Chicken Breast",         "6 oz marinated chicken breast, herb-seasoned",                    10.99),
    ("CHIX02", "Crispy Fried Chicken Sandwich",  "Hand-breaded chicken thigh on brioche bun with house slaw",       12.49),
    ("PSTA01", "Spaghetti Marinara",             "House-made marinara over al-dente spaghetti",                      9.99),
    ("PSTA02", "Fettuccine Alfredo",             "Fettuccine in cream parmesan sauce with fresh parsley",           11.99),
    ("SALAD01","Caesar Salad",                   "Romaine, house Caesar dressing, croutons, shaved Parmesan",        8.99),
    ("SALAD02","House Garden Salad",             "Mixed greens, cherry tomatoes, cucumber, carrots, choice of dressing", 7.49),
    ("SOUP01", "Soup of the Day",                "Chef's daily selection — ask your server",                         5.99),
    ("SOUP02", "French Onion Soup",              "Caramelized onion broth, toasted crouton, melted Gruyère",         7.99),
    ("APP01",  "Mozzarella Sticks",              "Hand-breaded mozzarella, served with marinara dipping sauce",      8.49),
    ("APP02",  "Chicken Wings (12 pc)",          "Choice of Buffalo, BBQ, or honey garlic; served with ranch",      14.99),
    ("APP03",  "Spinach Artichoke Dip",          "Warm dip served with toasted pita chips and tortilla chips",       9.99),
    ("PIZZA01","Margherita Pizza (12\")",         "Fresh mozzarella, basil, house tomato sauce",                    13.99),
    ("PIZZA02","Pepperoni Pizza (12\")",          "Classic pepperoni on house tomato sauce and mozzarella",          14.99),
    ("STEAK01","8 oz Sirloin Steak",             "USDA Choice sirloin, grilled to order, served with two sides",   24.99),
    ("FISH01", "Pan-Seared Salmon",              "6 oz Atlantic salmon, lemon butter caper sauce, seasonal veg",    18.99),
    ("FISH02", "Fish and Chips",                 "Beer-battered cod, seasoned fries, tartar sauce, coleslaw",       14.49),
    ("SAND01", "Turkey Club Sandwich",           "Sliced turkey, bacon, lettuce, tomato, avocado on sourdough",     11.99),
    ("SAND02", "Grilled Cheese Deluxe",          "Three-cheese blend on thick Texas toast, served with tomato soup", 9.49),
    ("SIDE01", "Seasoned French Fries",          "Crispy hand-cut fries with house seasoning blend",                 4.99),
    ("SIDE02", "Garlic Mashed Potatoes",         "Creamy mashed potatoes with roasted garlic and chive",             4.99),
    ("DSSRT01","New York Cheesecake",            "Rich NY-style cheesecake with berry compote",                      6.99),
    ("DSSRT02","Chocolate Lava Cake",            "Warm chocolate cake with molten center, vanilla ice cream",         7.49),
    ("BVGE01", "Fresh-Brewed Iced Tea",          "Unsweetened or sweetened, free refills",                           2.99),
]

wb = Workbook()
ws = wb.active
ws.title = "Items"

headers = ["ItemCode", "ItemName", "ItemDescription", "TypicalPrice"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font  = HEADER_FONT
    cell.fill  = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")

for r, (code, name, desc, price) in enumerate(items, 2):
    fill = ALT_FILL if r % 2 == 0 else None
    data = [code, name, desc, price]
    for c, v in enumerate(data, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = NORMAL_FONT
        if fill:
            cell.fill = fill
        cell.border = BORDER
        if c == 4:  # TypicalPrice
            cell.number_format = '"$"#,##0.00'

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 18

col_widths = [14, 36, 62, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = w

wb.save(r"C:\ClaudeOutput\Culinaire\SampleData\Items.xlsx")
print("Items.xlsx written.")
