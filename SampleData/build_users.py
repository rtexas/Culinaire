from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR_BG = "2B6B35"; HDR_FG = "FFFFFF"; ROW_ALT = "F0F7F0"; FN = "Arial"

def hdr(cell, h):
    cell.value = h
    cell.font = Font(name=FN, bold=True, color=HDR_FG, size=10)
    cell.fill = PatternFill("solid", start_color=HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def body(ws, r, row, n):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name=FN, size=10)
        cell.alignment = Alignment(vertical="center")
    if r % 2 == 0:
        for c in range(1, n+1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", start_color=ROW_ALT)

def autofit(ws, hdrs, rows):
    for i, h in enumerate(hdrs, 1):
        ml = len(str(h))
        for row in rows:
            ml = max(ml, len(str(row[i-1]) if i-1 < len(row) else ""))
        ws.column_dimensions[get_column_letter(i)].width = min(ml + 4, 40)

# Columns: Username, FullName, Email, RoleType, Password, IsActive,
#          Address1, Address2, Address3, City, State, PostalCode, Country
# Notes:
#   - Password: supplied for new users; blank = system default P@ssw0rd1! on import
#   - IsActive: true/false (omit or "true" = active)
#   - State: 2-letter code matching StatesRegions.Code
#   - Country: 2-letter code matching Countries.Code

users = [
    # Restaurant / Hospitality staff mix
    ("jsmith",     "John Smith",       "jsmith@culinaire.com",    "Administrator", "Admin#2024!",  "true",  "100 Main St",      "",           "", "Dallas",       "TX", "75201", "US"),
    ("mgarcia",    "Maria Garcia",     "mgarcia@culinaire.com",   "User",          "User#2024!",   "true",  "245 Oak Ave",      "Apt 3B",     "", "Austin",       "TX", "78701", "US"),
    ("dlee",       "David Lee",        "dlee@culinaire.com",      "User",          "User#2024!",   "true",  "810 Elm St",       "",           "", "Houston",      "TX", "77002", "US"),
    ("swilliams",  "Sarah Williams",   "swilliams@culinaire.com", "Viewer",        "View#2024!",   "true",  "33 Peach Tree Rd", "Suite 100",  "", "Atlanta",      "GA", "30303", "US"),
    ("rjohnson",   "Robert Johnson",   "rjohnson@culinaire.com",  "User",          "User#2024!",   "true",  "520 Michigan Ave", "",           "", "Chicago",      "IL", "60611", "US"),
    ("lmartinez",  "Laura Martinez",   "lmartinez@culinaire.com", "User",          "User#2024!",   "true",  "1200 Biscayne Blvd","",         "", "Miami",        "FL", "33132", "US"),
    ("kwilson",    "Kevin Wilson",     "kwilson@culinaire.com",   "Viewer",        "View#2024!",   "true",  "88 Bourbon St",    "",           "", "New Orleans",  "LA", "70116", "US"),
    ("aanderso",   "Amanda Anderson",  "aanderso@culinaire.com",  "User",          "User#2024!",   "true",  "400 Broadway",     "Fl 5",       "", "Nashville",    "TN", "37201", "US"),
    ("btaylor",    "Brian Taylor",     "btaylor@culinaire.com",   "User",          "User#2024!",   "true",  "900 Las Vegas Blvd","",         "", "Las Vegas",    "NV", "89101", "US"),
    ("cthomas",    "Catherine Thomas", "cthomas@culinaire.com",   "Viewer",        "View#2024!",   "true",  "215 Fremont St",   "",           "", "Las Vegas",    "NV", "89101", "US"),
    ("ejackson",   "Eric Jackson",     "ejackson@culinaire.com",  "User",          "User#2024!",   "true",  "1 Market Plaza",   "Ste 300",    "", "San Francisco","CA", "94105", "US"),
    ("nwhite",     "Nicole White",     "nwhite@culinaire.com",    "User",          "User#2024!",   "true",  "350 5th Ave",      "",           "", "New York",     "NY", "10118", "US"),
    ("pharris",    "Paul Harris",      "pharris@culinaire.com",   "Viewer",        "View#2024!",   "true",  "1600 Penn Ave NW", "",           "", "Washington",   "DC", "20006", "US"),
    ("tmartin",    "Tiffany Martin",   "tmartin@culinaire.com",   "User",          "User#2024!",   "true",  "200 E Colfax Ave", "",           "", "Denver",       "CO", "80203", "US"),
    ("gthompso",   "Greg Thompson",    "gthompso@culinaire.com",  "User",          "User#2024!",   "true",  "700 Pike St",      "",           "", "Seattle",      "WA", "98101", "US"),
    ("hgarciam",   "Hannah Garcia M.", "hgarciam@culinaire.com",  "Viewer",        "View#2024!",   "true",  "2121 N Harlem Ave","",          "", "Phoenix",      "AZ", "85001", "US"),
    ("jclark",     "James Clark",      "jclark@culinaire.com",    "User",          "User#2024!",   "true",  "500 W Madison St", "Apt 12",     "", "Chicago",      "IL", "60661", "US"),
    ("mlewis",     "Michelle Lewis",   "mlewis@culinaire.com",    "User",          "User#2024!",   "true",  "1 Bourbon St",     "",           "", "Louisville",   "KY", "40202", "US"),
    ("drobin",     "Daniel Robinson",  "drobin@culinaire.com",    "Viewer",        "View#2024!",   "true",  "600 S 4th St",     "Floor 2",    "", "Louisville",   "KY", "40202", "US"),
    ("awalker",    "Ashley Walker",    "awalker@culinaire.com",   "User",          "User#2024!",   "true",  "1500 S Main St",   "",           "", "Charlotte",    "NC", "28203", "US"),
    ("chall",      "Christopher Hall", "chall@culinaire.com",     "User",          "User#2024!",   "true",  "300 Riverwalk Pl", "",           "", "San Antonio",  "TX", "78205", "US"),
    ("syoung",     "Stephanie Young",  "syoung@culinaire.com",    "Viewer",        "View#2024!",   "true",  "80 Federal St",    "Ste 400",    "", "Boston",       "MA", "02110", "US"),
    ("bking",      "Brandon King",     "bking@culinaire.com",     "User",          "User#2024!",   "true",  "400 NE 2nd Ave",   "",           "", "Portland",     "OR", "97232", "US"),
    ("lwright",    "Lisa Wright",      "lwright@culinaire.com",   "User",          "User#2024!",   "true",  "1234 E Broad St",  "Unit B",     "", "Columbus",     "OH", "43205", "US"),
    ("mscott",     "Michael Scott",    "mscott@culinaire.com",    "Viewer",        "View#2024!",   "true",  "100 N Brand Blvd", "",           "", "Glendale",     "CA", "91203", "US"),
]

hdrs = ["Username","FullName","Email","RoleType","Password","IsActive",
        "Address1","Address2","Address3","City","State","PostalCode","Country"]

wb = Workbook(); ws = wb.active; ws.title = "Users"
ws.row_dimensions[1].height = 20

for c, h in enumerate(hdrs, 1): hdr(ws.cell(1, c), h)
for r, row in enumerate(users, 2): body(ws, r, row, len(hdrs))
autofit(ws, hdrs, users)
ws.freeze_panes = "A2"

path = r"C:\ClaudeOutput\Culinaire\SampleData\Users.xlsx"
wb.save(path)
print(f"Saved: {path}  ({len(users)} users)")
