from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR_BG = "2B6B35"; HDR_FG = "FFFFFF"; ROW_ALT = "F0F7F0"; FN = "Arial"

def hdr(cell, h):
    cell.value = h
    cell.font = Font(name=FN, bold=True, color=HDR_FG, size=10)
    cell.fill = PatternFill("solid", start_color=HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def body(ws, r, row, n):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name=FN, size=10)
        cell.alignment = Alignment(vertical="center")
    if r % 2 == 0:
        for c in range(1, n + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", start_color=ROW_ALT)

def autofit(ws, hdrs, rows):
    for i, h in enumerate(hdrs, 1):
        ml = len(str(h))
        for row in rows:
            ml = max(ml, len(str(row[i - 1]) if i - 1 < len(row) else ""))
        ws.column_dimensions[get_column_letter(i)].width = min(ml + 4, 50)

# Columns: Name, Code, Description
states = [
    ("Alabama",        "AL", "State in the southeastern United States"),
    ("Alaska",         "AK", "Northernmost state, located in the far northwest of North America"),
    ("Arizona",        "AZ", "State in the southwestern United States"),
    ("Arkansas",       "AR", "State in the south-central United States"),
    ("California",     "CA", "Most populous state, located on the Pacific Coast"),
    ("Colorado",       "CO", "State in the Mountain West region"),
    ("Connecticut",    "CT", "State in the New England region"),
    ("Delaware",       "DE", "Smallest state by area in the Mid-Atlantic region"),
    ("Florida",        "FL", "Southernmost contiguous state, located on a peninsula"),
    ("Georgia",        "GA", "State in the southeastern United States"),
    ("Hawaii",         "HI", "Island state located in the Pacific Ocean"),
    ("Idaho",          "ID", "State in the Pacific Northwest region"),
    ("Illinois",       "IL", "State in the Midwest, home to Chicago"),
    ("Indiana",        "IN", "State in the Midwest region"),
    ("Iowa",           "IA", "State in the Midwest known for agriculture"),
    ("Kansas",         "KS", "State in the Great Plains region"),
    ("Kentucky",       "KY", "State in the southeastern United States"),
    ("Louisiana",      "LA", "State in the Deep South known for Creole culture"),
    ("Maine",          "ME", "Northeasternmost state in New England"),
    ("Maryland",       "MD", "State in the Mid-Atlantic region"),
    ("Massachusetts",  "MA", "State in the New England region"),
    ("Michigan",       "MI", "State in the Great Lakes region"),
    ("Minnesota",      "MN", "State in the upper Midwest"),
    ("Mississippi",    "MS", "State in the Deep South region"),
    ("Missouri",       "MO", "State in the Midwest at the confluence of the Missouri and Mississippi rivers"),
    ("Montana",        "MT", "State in the northwestern United States known for mountains"),
    ("Nebraska",       "NE", "State in the Great Plains region"),
    ("Nevada",         "NV", "State in the western United States known for Las Vegas"),
    ("New Hampshire",  "NH", "State in the New England region"),
    ("New Jersey",     "NJ", "State in the Mid-Atlantic region"),
    ("New Mexico",     "NM", "State in the southwestern United States"),
    ("New York",       "NY", "State in the northeastern United States, home to New York City"),
    ("North Carolina", "NC", "State in the southeastern United States"),
    ("North Dakota",   "ND", "State in the upper Midwest and Great Plains"),
    ("Ohio",           "OH", "State in the Midwest region"),
    ("Oklahoma",       "OK", "State in the south-central United States"),
    ("Oregon",         "OR", "State in the Pacific Northwest region"),
    ("Pennsylvania",   "PA", "State in the northeastern United States"),
    ("Rhode Island",   "RI", "Smallest state by area, located in New England"),
    ("South Carolina", "SC", "State in the southeastern United States"),
    ("South Dakota",   "SD", "State in the upper Midwest and Great Plains"),
    ("Tennessee",      "TN", "State in the southeastern United States"),
    ("Texas",          "TX", "Second largest state, located in the south-central United States"),
    ("Utah",           "UT", "State in the Mountain West region"),
    ("Vermont",        "VT", "State in the New England region"),
    ("Virginia",       "VA", "State in the Mid-Atlantic and southeastern United States"),
    ("Washington",        "WA", "State in the Pacific Northwest region"),
    ("Washington DC",     "DC", "Federal district and capital of the United States"),
    ("West Virginia",  "WV", "State in the Appalachian region"),
    ("Wisconsin",      "WI", "State in the upper Midwest region"),
    ("Wyoming",        "WY", "State in the Mountain West, least populous state"),
]

hdrs = ["Name", "Code", "Description"]

wb = Workbook(); ws = wb.active; ws.title = "States"
ws.row_dimensions[1].height = 20

for c, h in enumerate(hdrs, 1): hdr(ws.cell(1, c), h)
for r, row in enumerate(states, 2): body(ws, r, row, len(hdrs))
autofit(ws, hdrs, states)
ws.freeze_panes = "A2"

path = r"C:\ClaudeOutput\Culinaire\SampleData\States.xlsx"
wb.save(path)
print(f"Saved: {path}  ({len(states)} states)")
