import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FONT = Font(bold=True, name='Arial', size=10)
HEADER_FILL = PatternFill('solid', start_color='2B6B35', end_color='2B6B35')
HEADER_FONT_WHITE = Font(bold=True, name='Arial', size=10, color='FFFFFF')
DATA_FONT = Font(name='Arial', size=10)

def style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 14)

def style_row(ws, row, data):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row, col, val)
        cell.font = DATA_FONT

# ── Job Roles ────────────────────────────────────────────────────────────────

job_roles = [
    ("Salaried_GM",     "General Manager",       "Oversees all restaurant operations",                    "Exempt",  "1"),
    ("Salaried_AM",     "Assistant Manager",      "Supports GM and manages day-to-day operations",         "Exempt",  "1"),
    ("Hourly_SL",       "Shift Leader",           "Leads team during assigned shift",                      "Hourly",  "1"),
    ("Hourly_Server",   "Server",                 "Takes orders and serves food and beverages",            "Hourly",  "1"),
    ("Hourly_Cook",     "Line Cook",              "Prepares food items per menu standards",                "Hourly",  "1"),
    ("Hourly_Prep",     "Prep Cook",              "Handles food preparation and mise en place",            "Hourly",  "1"),
    ("Hourly_Host",     "Host/Hostess",           "Greets and seats guests, manages waitlist",             "Hourly",  "1"),
    ("Hourly_Cashier",  "Cashier",                "Processes payments and operates POS system",            "Hourly",  "1"),
    ("Hourly_Busser",   "Busser",                 "Clears and resets tables, supports service team",       "Hourly",  "1"),
    ("Hourly_Dishwash", "Dishwasher",             "Maintains cleanliness of kitchen equipment and dishes", "Hourly",  "1"),
]

wb_jobs = openpyxl.Workbook()
ws_jobs = wb_jobs.active
ws_jobs.title = "JobRoles"
headers_jobs = ["ExternalID", "Name", "Description", "PayType", "IsActive"]
style_header(ws_jobs, headers_jobs)
for i, row in enumerate(job_roles, 2):
    style_row(ws_jobs, i, row)
ws_jobs.column_dimensions['B'].width = 22
ws_jobs.column_dimensions['C'].width = 52
wb_jobs.save(r'C:\ClaudeOutput\Culinaire\SampleData\JobRoles.xlsx')
print("JobRoles.xlsx created")

# ── Employees ────────────────────────────────────────────────────────────────
# Locations from Locations.xlsx: 001=Main Restaurant, 002=Bar & Lounge,
#                                003=Catering, 004=Delivery
# Multiple rows per ExternalID = multiple location assignments

employees_single = [
    # (ExternalID, Name, Description, IsActive, LocationCode)
    # Managers — span multiple locations
    ("EMP001", "James Martinez",    "General Manager",    "1", "001"),
    ("EMP001", "James Martinez",    "General Manager",    "1", "002"),
    ("EMP002", "Sarah Johnson",     "Assistant Manager",  "1", "001"),
    ("EMP002", "Sarah Johnson",     "Assistant Manager",  "1", "003"),
    ("EMP003", "Michael Chen",      "Shift Leader",       "1", "001"),
    ("EMP003", "Michael Chen",      "Shift Leader",       "1", "002"),
    ("EMP004", "Emily Rodriguez",   "Shift Leader",       "1", "002"),
    ("EMP005", "David Kim",         "Shift Leader",       "1", "003"),
    ("EMP005", "David Kim",         "Shift Leader",       "1", "004"),
    ("EMP006", "Lisa Thompson",     "Assistant Manager",  "1", "002"),
    ("EMP006", "Lisa Thompson",     "Assistant Manager",  "1", "003"),
    ("EMP007", "Robert Williams",   "General Manager",    "1", "003"),
    ("EMP007", "Robert Williams",   "General Manager",    "1", "004"),
    # Servers — Main Restaurant & Bar & Lounge
    ("EMP008", "Amanda Davis",      "Server",             "1", "001"),
    ("EMP009", "Christopher Lee",   "Server",             "1", "001"),
    ("EMP010", "Jessica Brown",     "Server",             "1", "001"),
    ("EMP011", "Matthew Wilson",    "Server",             "1", "002"),
    ("EMP012", "Ashley Taylor",     "Server",             "1", "002"),
    ("EMP013", "Daniel Anderson",   "Server",             "1", "002"),
    ("EMP014", "Stephanie Thomas",  "Server",             "1", "003"),
    ("EMP015", "Joshua Jackson",    "Server",             "1", "003"),
    # Line Cooks — all four locations
    ("EMP016", "Brittany White",    "Line Cook",          "1", "001"),
    ("EMP017", "Andrew Harris",     "Line Cook",          "1", "001"),
    ("EMP018", "Megan Martin",      "Line Cook",          "1", "001"),
    ("EMP019", "Ryan Garcia",       "Line Cook",          "1", "002"),
    ("EMP020", "Kayla Martinez",    "Line Cook",          "1", "002"),
    ("EMP021", "Brandon Robinson",  "Line Cook",          "1", "003"),
    ("EMP022", "Samantha Clark",    "Line Cook",          "1", "003"),
    ("EMP023", "Tyler Rodriguez",   "Line Cook",          "1", "004"),
    # Prep Cooks
    ("EMP024", "Rachel Lewis",      "Prep Cook",          "1", "001"),
    ("EMP025", "Nathan Lee",        "Prep Cook",          "1", "001"),
    ("EMP026", "Lauren Walker",     "Prep Cook",          "1", "002"),
    ("EMP027", "Justin Hall",       "Prep Cook",          "1", "003"),
    ("EMP028", "Amber Allen",       "Prep Cook",          "1", "004"),
    # Hosts — Main Restaurant & Bar & Lounge
    ("EMP029", "Kevin Young",       "Host/Hostess",       "1", "001"),
    ("EMP030", "Hannah Hernandez",  "Host/Hostess",       "1", "001"),
    ("EMP031", "Aaron King",        "Host/Hostess",       "1", "002"),
    # Cashiers — all locations
    ("EMP032", "Tiffany Wright",    "Cashier",            "1", "001"),
    ("EMP033", "Eric Lopez",        "Cashier",            "1", "002"),
    ("EMP034", "Heather Hill",      "Cashier",            "1", "003"),
    ("EMP035", "Adam Scott",        "Cashier",            "1", "004"),
    # Bussers
    ("EMP036", "Jennifer Green",    "Busser",             "1", "001"),
    ("EMP037", "Steven Adams",      "Busser",             "1", "001"),
    ("EMP038", "Melissa Baker",     "Busser",             "1", "002"),
    ("EMP039", "Jonathan Gonzalez", "Busser",             "1", "002"),
    ("EMP040", "Crystal Nelson",    "Busser",             "1", "003"),
    # Dishwashers
    ("EMP041", "Patrick Carter",    "Dishwasher",         "1", "001"),
    ("EMP042", "Danielle Mitchell", "Dishwasher",         "1", "001"),
    ("EMP043", "Timothy Perez",     "Dishwasher",         "1", "002"),
    ("EMP044", "Courtney Roberts",  "Dishwasher",         "1", "003"),
    ("EMP045", "Jeremy Turner",     "Dishwasher",         "1", "004"),
    # Additional staff
    ("EMP046", "Vanessa Phillips",  "Shift Leader",       "1", "001"),
    ("EMP047", "Charles Campbell",  "Server",             "1", "001"),
    ("EMP047", "Charles Campbell",  "Server",             "1", "002"),
    ("EMP048", "Monica Parker",     "Line Cook",          "1", "002"),
    ("EMP049", "Raymond Evans",     "Server",             "1", "003"),
    ("EMP049", "Raymond Evans",     "Server",             "1", "004"),
    ("EMP050", "Jasmine Edwards",   "Cashier",            "1", "001"),
    ("EMP050", "Jasmine Edwards",   "Cashier",            "1", "003"),
]

wb_emp = openpyxl.Workbook()
ws_emp = wb_emp.active
ws_emp.title = "Employees"
headers_emp = ["ExternalID", "Name", "Description", "IsActive", "LocationCode"]
style_header(ws_emp, headers_emp)
for i, row in enumerate(employees_single, 2):
    style_row(ws_emp, i, row)
ws_emp.column_dimensions['A'].width = 12
ws_emp.column_dimensions['B'].width = 24
ws_emp.column_dimensions['C'].width = 22
ws_emp.column_dimensions['D'].width = 10
ws_emp.column_dimensions['E'].width = 14
wb_emp.save(r'C:\ClaudeOutput\Culinaire\SampleData\Employees.xlsx')
print("Employees.xlsx created")
print("Done.")
