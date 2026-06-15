import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# (Name, Multiplier, Used in EOD Sales, Used in EOD Graph, Description)
sections = [
    ("Sales",         1,  1, 1, "Revenue items — added to the EOD total"),
    ("Dispositions", -1,  1, 0, "Comps, voids, discounts, and other reductions — subtracted from EOD Sales Total"),
    ("Liabilities",  -1,  0, 0, "Expense and liability items — displayed for reference, not included in the Grand Total"),
    ("Unit Accounts", 1,  0, 0, "Summary Accounts Used"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EOD Sections"

HDR_BG = "2B6B35"; HDR_FG = "FFFFFF"; ALT = "F0F7F0"
header_fill  = PatternFill("solid", fgColor=HDR_BG)
header_font  = Font(name="Arial", bold=True, color=HDR_FG, size=11)
header_align = Alignment(horizontal="center", vertical="center")
alt_fill     = PatternFill("solid", fgColor=ALT)
data_font    = Font(name="Arial", size=10)
thin         = Border(
    left=Side(style="thin", color="C8DFC8"), right=Side(style="thin", color="C8DFC8"),
    top=Side(style="thin", color="C8DFC8"),  bottom=Side(style="thin", color="C8DFC8"),
)

headers    = ["Name", "Multiplier", "Used in EOD Sales", "Used in EOD Graph", "Description"]
col_widths = [18,     12,           20,                  20,                  55]

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = header_align; cell.border = thin
    ws.column_dimensions[cell.column_letter].width = w
ws.row_dimensions[1].height = 20

for row_idx, (name, mult, use_in_eod, use_in_graph, desc) in enumerate(sections, start=2):
    fill = alt_fill if row_idx % 2 == 0 else PatternFill()
    for col, val in enumerate([name, mult, use_in_eod, use_in_graph, desc], start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = data_font; cell.fill = fill; cell.border = thin
        cell.alignment = Alignment(vertical="center",
                                   horizontal="center" if col in (2, 3, 4) else "left")

ws.freeze_panes = "A2"
out = r"C:\ClaudeOutput\Culinaire\SampleData\EOD Sections.xlsx"
wb.save(out)
print(f"Saved: {out}  ({len(sections)} sections)")
